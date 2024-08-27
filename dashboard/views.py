from django.template.defaultfilters import striptags
import psycopg2
import requests,json
from django.shortcuts import render
from dashboard.models import *
from survey.models import *
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.db import connection
from application_master.models import Boundary
from ast import literal_eval
from collections import OrderedDict
import colorsys
import logging
from dateutil.relativedelta import relativedelta
from django.shortcuts import render,HttpResponse
from django.db import connection
from dashboard.models import DashboardSummaryLog,MonthlyDashboard,Remarks,STATUS_CHOICES,ArrayField,ChartMeta
from application_master.models import UserPartnerMapping,UserProjectMapping,PartnerMissionMapping
from datetime import datetime,date, timedelta
from mis.models import *
from survey.views import get_pagination,JsonAnswer
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db.models import Case, Value, When, Q , Sum
from mfv_mis.settings import DASHBOARD_SUBMISSION_DAY
from uuid import uuid4
logger = logging.getLogger(__name__)
from rest_framework.generics import CreateAPIView
from rest_framework.response import Response
from reports.views import load_user_details
# ****************************************************************************
# Login Function
# ****************************************************************************

def filter_location_data(request):
    # zone_list,state_list,partner_list,donor_list=[],[],[],[]
    partner_filter,zone_filter,state_filter,district_filter,donor_filter = None,None,None,None,None
    selected_items = []
    user_filter_data = []
    user_role = request.session.get('role_id')
    if request.session.get('role_id')in (8,9,10):
        get_state_partner_linkage = ApplicationUserStateLinkage.objects.filter(status=2,user_id=request.user.id).values_list('state',flat=True)
        zone_id = State.objects.filter(status=2,id__in=get_state_partner_linkage).values_list('zone',flat=True)
        zone_filter = Zone.objects.filter(status=2,id__in=zone_id).order_by('name')
        selected_items.insert(0, str(zone_id[0]))
        user_filter_data.insert(0, str(zone_id[0]))
        state_filter = State.objects.filter(status=2,id__in=get_state_partner_linkage,zone__in=zone_id).order_by('name')
        partner_id = Partner.objects.filter(status=2,state_id__in=get_state_partner_linkage).values_list('id',flat=True)
        partner_filter = Partner.objects.filter(status=2,id__in=partner_id).order_by('name')
        donor_id = DonorPartnerLinkage.objects.filter(status=2,partner_id__in=partner_id).values_list('donor_id',flat=True)
        donor_filter = Donor.objects.filter(status=2,id__in=donor_id).order_by('name')
        if request.POST.get('state', '') != '':
            state_id = request.POST.get('state', '')
            district_filter = District.objects.filter(status=2,state_id=int(state_id)).order_by('name')
            selected_items.insert(1,'['+str(request.POST.get('state', ''))+']')
            user_filter_data.insert(1,request.POST.get('state', ''))
        else:
            selected_items.insert(1, str([i for i in get_state_partner_linkage]))
            user_filter_data.insert(1,request.POST.get('state', ''))
        # selected_items.insert(2,request.POST.get('district', ''))
        # user_filter_data.insert(2,request.POST.get('district', ''))
        if request.POST.get('partner', '') != '':
            selected_items.insert(3,'['+str(request.POST.get('partner', ''))+']')
            user_filter_data.insert(3, request.POST.get('partner', ''))
        else:
            selected_items.insert(3,request.POST.get('partner',str([i for i in partner_id])))
            user_filter_data.insert(3, request.POST.get('partner', ''))
        # selected_items.insert(4,request.POST.get('donor', ''))
        # user_filter_data.insert(4,request.POST.get('donor', ''))
        # selected_items.insert(2,request.POST.get('district', ''))
        # user_filter_data.insert(2,request.POST.get('district', '')) 
    elif request.session.get('role_id') == 4:
        partner_users = UserPartnerLinkage.objects.get(user_id=request.user.id)
        partner_filter=Partner.objects.filter(status=2,id=partner_users.partner_id).order_by('name')
        state_id = Partner.objects.filter(status=2,id=partner_users.partner_id).values_list('state_id',flat=True)
        zone_id = State.objects.filter(status=2,id__in=state_id).values_list('zone',flat=True)
        zone_filter = Zone.objects.filter(status=2,id__in=zone_id).order_by('name')
        if request.POST.get('zone', '') != '':
            state_filter = State.objects.filter(status=2,id__in=state_id).order_by('name')
            selected_items.insert(0, request.POST.get('zone', ''))
            user_filter_data.insert(0, request.POST.get('zone', ''))
        else:
            selected_items.insert(0, request.POST.get('zone', ''))
            user_filter_data.insert(0, request.POST.get('zone', ''))
        if request.POST.get('state', '') != '':
            district_filter = District.objects.filter(status=2,state__in = state_id).order_by('name')
            selected_items.insert(1,'['+str(request.POST.get('state', ''))+']')
            user_filter_data.insert(1,request.POST.get('state', ''))
        else:
            selected_items.insert(1,request.POST.get('state', ''))
            user_filter_data.insert(1,request.POST.get('state', ''))
        # selected_items.insert(2,request.POST.get('district', ''))
        # user_filter_data.insert(2,request.POST.get('district', ''))
        selected_items.insert(3,'['+str(partner_users.partner_id)+']')
        user_filter_data.insert(3,str(partner_users.partner_id))
        donor_id = DonorPartnerLinkage.objects.filter(status=2,partner_id=partner_users.partner_id).values_list('donor_id',flat=True)
        donor_filter = Donor.objects.filter(status=2,id__in=donor_id).order_by('name')
        # selected_items.insert(4,request.POST.get('donor', ''))
        # user_filter_data.insert(4,request.POST.get('donor', ''))
    else:
        zone_filter = Zone.objects.filter(status=2).order_by('name')
        if request.POST.get('zone', '') != '':
            zone_id = request.POST.get('zone', '')
            state_filter = State.objects.filter(status=2,zone_id=int(zone_id)).order_by('name')
            selected_items.insert(0, request.POST.get('zone', ''))
            user_filter_data.insert(0, request.POST.get('zone', ''))
        else:
            selected_items.insert(0, request.POST.get('zone', ''))
            user_filter_data.insert(0, request.POST.get('zone', ''))
        if request.POST.get('state', '') != '':
            state_id = request.POST.get('state', '')
            district_filter = District.objects.filter(status=2,state_id=int(state_id)).order_by('name')
            selected_items.insert(1,'['+str(request.POST.get('state', ''))+']') 
            user_filter_data.insert(1,request.POST.get('state', ''))
        else:
            selected_items.insert(1,request.POST.get('state', ''))
            user_filter_data.insert(1,request.POST.get('state', ''))
        # selected_items.insert(2,request.POST.get('district', ''))
        # user_filter_data.insert(2,request.POST.get('district', ''))
        partner_filter=Partner.objects.filter(status=2).order_by('name')
        if request.POST.get('partner', '') != '':
            selected_items.insert(3,'['+str(request.POST.get('partner', ''))+']')
            user_filter_data.insert(3, request.POST.get('partner', ''))
        else:
            selected_items.insert(3, request.POST.get('partner', '')) 
            user_filter_data.insert(3, request.POST.get('partner', ''))
        donor_filter=Donor.objects.filter(status=2).order_by('name')
        # selected_items.insert(4,request.POST.get('donor', ''))
        # user_filter_data.insert(4,request.POST.get('donor', ''))
    selected_items.insert(4,request.POST.get('donor', ''))
    user_filter_data.insert(4,request.POST.get('donor', ''))
    selected_items.insert(2,request.POST.get('district', ''))
    user_filter_data.insert(2,request.POST.get('district', ''))
    from_date = request.POST.get('from_date', '') 
    to_date = request.POST.get('to_date', '')
    selected_items.insert(5,from_date)
    user_filter_data.insert(5,from_date)
    selected_items.insert(6,to_date)  
    user_filter_data.insert(6,to_date)
    return zone_filter,partner_filter,state_filter,district_filter,donor_filter,selected_items,user_filter_data,user_role

# ****************************************************************************
# Function to load user details to session
# ****************************************************************************


def load_user_details_to_sessions(request):
    # Getting the user role config if not it will raise exception
    try:
        user_role_location_level_config = UserRoleLocationLevelConfig.objects.get(
            user=request.user, status=2)
    except UserRoleLocationLevelConfig.DoesNotExist:
        # user_role_location_level_config = None
        configure_error = 'Username not configured . Please contact administrator.'
        return configure_error

    # User group for checking group permission with menu permission
    user_group = user_role_location_level_config.group
    location_hierarchy_type_id = user_role_location_level_config.location_hierarchy_type.id

    # Return user location relation id and object id list , if user mapped with multiple location.
    user_object_id = UserLocationRelation.objects.filter(
        UserRoleLocationLevelConfig=user_role_location_level_config, content_type__id=location_hierarchy_type_id, status=2).values_list('object_id', flat=True)

    menus = Menu.objects.filter(status=2)
    menu_to_display = []

    for menu in menus:
        if user_group.permissions.filter(id=menu.model_permission.id).exists():
            menu_to_display.append(
                (menu.name, menu.slug, menu.icon, menu.feature_link))

    request.session['menus'] = menu_to_display
    # request.session['user_group_id'] = user_group.id
    request.session['location_hierarchy_type_id'] = location_hierarchy_type_id
    request.session['user_object_id'] = list(
        user_object_id)

# #****************************************************************************
# # update stackabar chart data to replace dummy data with actual values
# #****************************************************************************


def set_dynamic_column_stack_chart_data(sql, headers):
    cursor = None
    try:
        chart_row_id_value = []
        cursor = connection.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
        newdata = []
        newdata.append(headers)
        for row in rows:
            row_data = list(row)
            chart_id_val = row_data.pop(0)
            chart_row_id_value.append(chart_id_val)
            newdata.append(row_data)
            #new_row_data = []
    finally:
        if cursor:
            cursor.close()
    return newdata, chart_row_id_value


# #****************************************************************************
# # update pie chart data to replace dummy data with actual values
# #****************************************************************************


def set_pie_chart_data(sql, labels=None):
    cursor = connection.cursor()
    cursor.execute(sql)
    descr = cursor.description
    rows = cursor.fetchall()
    data = []
    if labels:
        data = [dict(zip([column for column in labels], row))for row in rows]
        return data[0].items()
    else:
        return rows


def set_card_chart_data(sql, labels=None):
    cursor = connection.cursor()
    cursor.execute(sql)
    descr = cursor.description
    rows = cursor.fetchall()
    data = []
    if labels:
        data =[(labels[idx],row[0]) for idx,row in enumerate(rows)]
        return data
    else:
        return rows

# #****************************************************************************
# # update column chart data  and labels to replace dummy data with actual values
# # labels only for dynamic bars - last 6 months kind of charts
# #****************************************************************************


def set_column_chart_data(sql, labels):
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
        descr = cursor.description
        counter = 0
        data = []
        if rows:
            # data = [dict(zip([column for column in labels], row))for row in rows]
            # row = rows[0]
            for i in range(0,len(rows)):
                value = (labels[i],rows[i][1])
                data.append(value)

            return data
        else:
            data = []
            row = None
    finally:
        if cursor:
            cursor.close()
    return data[0].items()



def set_bar_chart_dynamic_lable(sql):
    cursor = None
    try:
        # query output strucutre - "location name", location_id, data values
        # location_id not required in chart_data so remove and add it to dict 
        cursor = connection.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
        #descr = cursor.description
        data = []
        chart_row_id_value = []
        for idx,i in enumerate(rows):
            i = list(i)
            #i[1] = int(i[1])
            loc_id = i.pop(0)
            chart_row_id_value.append(int(loc_id))
            data.append(i)
    finally:
        if cursor:
            cursor.close()
    return data, chart_row_id_value

def table_html_chart(sql):
    cursor = None
    cursor = connection.cursor()
    cursor.execute(sql)
    rows = cursor.fetchall()
    return rows

def custom_table_html_chart(sql):
    cursor = None
    cursor = connection.cursor()
    cursor.execute(sql)
    rows = cursor.fetchall()
    rows = [json.loads(data) for data in rows[0]]
    return rows

def custom_header_html_chart(sql):
    cursor = None
    cursor = connection.cursor()
    cursor.execute(sql)
    rows = cursor.fetchall()
    rows = [json.loads(rows[0][0])]
    return rows

def dynamic_table_html_chart(sql):
    cursor = None
    cursor = connection.cursor()
    cursor.execute(sql)
    rows = cursor.fetchall()
    return rows

# {'sql_query': 'select mc.name as colony_name, coalesce(approved_count,0) as approved_count from masterdata_colony as mc left outer join dash_board_benefits_enrollment_view dbev on dbev.colony_id = mc.id order by trim(lower(mc.name))', 'col_headers': ['Colony Name', 'Board Benefits']}
def get_location_data(request,user_id):
    state_list,district_list,block_list,survey_list,project_list = [],[],[],[],[]
    location_data = load_user_details_to_session(request)
    if location_data[0][0] != 0:
        user =  User.objects.get(id= user_id)
        # if user is not None:
        #     if not user.is_superuser and not user.is_staff:
        user_boundary_levelcode = UserRoles.objects.get(user=user_id)
        project_list = OrganizationLocation.objects.filter(user_id=user_boundary_levelcode.id).exclude(project_id=None,active=0).values_list('project',flat=True)
        projectlist = Project.objects.filter(id__in=ProjectUserRelation.objects.filter(user=user.userroles).exclude(active=0)).values_list("id",flat=True)
        project_list = list(set(list(project_list)+list(projectlist)))
        survey = Lineitem.objects.filter(project__in = project_list).exclude(active=0).values_list('activity', flat=True)
        survey_list = list(set(survey))
        if None in survey_list:
            survey_list.remove(None)
        boundary_object = Boundary.objects.get_or_none(id=user_boundary_levelcode.get_location_type()[0])
        state_list = [i[0] for i in location_data[0][2]]
        district_list = [i[0] for i in location_data[1][2]]
        block_list = [i[0] for i in location_data[2][2]]
    return state_list,district_list,block_list,survey_list,project_list


def replace_filter_values(sql_query,selected_items):
    f_start_value = str(selected_items[1]) + '-04'
    f_end_value = selected_items[0]
    sql_query = sql_query.replace('@@fy_start_value', f_start_value)
    sql_query = sql_query.replace('@@fy_end_value', f_end_value)
    return sql_query


def filter_conditions(request, sql_query, filter_info,selected_items):
    user_id = request.user.id
    district_list,partner_list = [],[]
    # district_list,project_list = get_location_data(request,user_id)
    filter_cond = filter_info.get('filter_cond','')
    #logger.error("selected_items:" + str(selected_items))
    fin_year_cond,quarterly_year_cond,partner_cond,dist_cond,block_cond = "","","","",""
    for key in filter_info['filter_cond'].keys():
        if key == 'financial_year':
            if selected_items[0] != '':
                fin_year_cond = filter_info['filter_cond'][key].replace('@@filter_value',f"'{selected_items[0]}'")
        elif key == 'quarterly_year':
            if selected_items[1] != '':
                quarterly_year_cond = filter_info['filter_cond'][key].replace('@@filter_value',f"'{selected_items[1]}'")
        elif key == 'partner':
            if selected_items[2] != '':
                partner_cond = filter_info['filter_cond'][key].replace('@@filter_value',f"{selected_items[2]}")
            elif partner_list:
                partner_cond = filter_info['filter_cond'][key].replace('@@filter_value',f"{str(parter_list)[1:-1]}")
        elif key == 'district':
            if selected_items[3] != '':
                dist_cond = filter_info['filter_cond'][key].replace('@@filter_value',f"{selected_items[3]}")
            elif district_list:
                dist_cond = filter_info['filter_cond'][key].replace('@@filter_value',f"{str(district_list)[1:-1]}")
    sql_query = sql_query.replace('@@financial_year_filter',fin_year_cond).replace('@@quarterly_year_filter',quarterly_year_cond).replace('@@partner_filter',partner_cond).replace('@@district_filter',dist_cond).replace('@@block_filter',block_cond)
    #logger.error("sql_query:" + sql_query)
    return sql_query

# #****************************************************************************
# # update table chart data to replace dummy data with actual values
# #****************************************************************************

def set_dynamic_table_chart_data(sql, headers):
    cursor = connection.cursor()
    cursor.execute(sql)
    rows = cursor.fetchall()
    newdata = []
    chart_row_id_value = []
    newdata.append(headers)
    data = []
    data.insert(0,tuple(headers))
    for idx,row in enumerate(rows):
        # newdata.append(list(map(str, list(row))))
       data.insert(idx+1,row)
    return data


# #****************************************************************************
# # Get Financial year
# #****************************************************************************

def getfinancialyear(fin_yaer):
    today = datetime.today()
    current_year = today.year
    diff_finyear = current_year - int(fin_yaer.split('-')[0])
    financial_year = []
    financial_year.append(fin_yaer)
    for year in range(diff_finyear):
        if int(financial_year[year].split('-')[1]) != current_year:
            financial_year.append(f"{financial_year[year].split('-')[1]}-{int(financial_year[year].split('-')[1])+1}")
        else:
            if today.month >= 4:
                financial_year.append(f"{financial_year[year].split('-')[1]}-{int(financial_year[year].split('-')[1])+1}")
    return financial_year


def getacademicyear(fin_yaer):
    today = datetime.today()
    current_year = today.year
    diff_finyear = current_year - int(fin_yaer.split('-')[0])
    financial_year = []
    financial_year.append(fin_yaer)
    for year in range(diff_finyear):
        if int(financial_year[year].split('-')[1]) != current_year:
            financial_year.append(f"{financial_year[year].split('-')[1]}-{int(financial_year[year].split('-')[1])+1}")
        else:
            if today.month >= 4:
                financial_year.append(f"{financial_year[year].split('-')[1]}-{int(financial_year[year].split('-')[1])+1}")
    return financial_year

#Header Dynmic

def custom_header_query(sql_query):
    cursor = connection.cursor()
    cursor.execute(sql_query)
    rows = cursor.fetchall()
    headers = json.loads(rows[0][0])
    return headers

# #****************************************************************************
# # Dashboard
# #****************************************************************************

def gethalfyear(financial_year):
    selected_fy = financial_year
    selected_year = selected_fy.split('-')
    today = datetime.today()
    result_set = []
    current_year = today.year
    current_month = today.month
    if current_year == int(selected_year[0]) and current_month >= 4 and current_month <= 6:
        result_set.append('Q1')
    elif current_year == int(selected_year[0]) and current_month >= 7 and current_month <= 9:
        result_set.append('Q1')
        result_set.append('Q2')
    elif current_year == int(selected_year[0]) and current_month >= 10 and current_month <= 12:
        result_set.append('Q1')
        result_set.append('Q2')
        result_set.append('Q3')
    elif current_year == int(selected_year[1]) and current_month >= 1 and current_month <= 3:
        result_set.append('Q1')
        result_set.append('Q2')
        result_set.append('Q3')
        result_set.append('Q4')
    return result_set

def user_projects(user):
    projectlist = Project.objects.filter(active=2).values('id','name').order_by('name')
    return projectlist

def get_district_list(selected_partner_list):
    partner_mission_id = PartnerMissionMapping.objects.filter(active=2,partner__in=selected_partner_list).values_list('id',flat=True)
    project_id = Project.objects.filter(active=2,partner_mission_mapping__in=partner_mission_id).values_list('district',flat=True)
    master_district_ids = Project.objects.filter(active=2,id__in=project_id).values_list('district_id',flat=True)
    boundarys = Boundary.objects.filter(active=2,code__in=[int(dist_id) for dist_id in master_district_ids]).values('id','name').order_by('name')
    return boundarys

class PartnerDistricts(CreateAPIView):
    def post(self, request):
        data = request.data
        partner_id = data.getlist('partner_id[]') or data.getlist('partner_id')
        partner_mission_id = PartnerMissionMapping.objects.filter(active=2,partner__in=partner_id).values_list('id',flat=True)
        project_id = Project.objects.filter(active=2,partner_mission_mapping__in=partner_mission_id).values_list('district',flat=True)
        master_district_ids = Project.objects.filter(active=2,id__in=project_id).values_list('district_id',flat=True)
        boundarys = Boundary.objects.filter(active=2,code__in=[int(dist_id) for dist_id in master_district_ids]).order_by('name')
        print(boundarys)
        data = [{'id': i.id, 'name': "{}".format(i.name) if i.name else i.code} for i in boundarys]
        print(data)
        response = ({'status': 2, 'data': data, 'message': 'success', })
        return Response(response)


@login_required(login_url='/login/')
def dashboard(request):
    heading = 'Dashboard'
    page_slug = request.GET.get('page_slug')
    financial_year_list = getfinancialyear('2024-2025')
    selected_financial_year = request.POST.get('finacial-year',financial_year_list[0])
    quarterly_year_list = gethalfyear(selected_financial_year)
    selected_quarterly_year = request.POST.get('quarterly-year',quarterly_year_list[-1])
    selected_partner_list = request.POST.getlist('Partner',[])
    selected_district_list = request.POST.getlist('District',[])
    partner_list_id = ','.join(str(eval(i)) for i in request.POST.getlist('Partner'))
    district_list_id = ','.join(str(eval(i)) for i in request.POST.getlist('District'))
    print(selected_quarterly_year,selected_financial_year,'-------------')
    user_details = load_user_details(request)
    if 'user_partner_list' in request.session:
        partner_ids = request.session['user_partner_list']
    partnerlist = Partner.objects.filter(id__in=partner_ids).values("id","name")
    if 'user_district_list' in request.session:
        district_ids = request.session['user_district_list']
    district_list = []
    if selected_partner_list:
        district_list = get_district_list(selected_partner_list)
    try:
        cht = ChartMeta.objects.filter(
            active=2,page_slug=page_slug).order_by('display_order')
        chart_list = []
        # financial_year = getfinancialyear(request.POST.get('from_month',datetime.today().strftime("%Y-%m")))
        # index 5,6,7 are chcl_district_id and chcl_block_id,chcl_fy are used to pass as query parameters when dispalying dashboard on user click action
        selected_items =[selected_financial_year,selected_quarterly_year,partner_list_id,district_list_id]
        for i in cht:
            if i.chart_type == 8:
                # chart_type values: 
                # 1=Column Chart, 2=Pie Chart, 3=Table Chart, 4=Bar Chart, 5=Column Stack, 6=Bar Dynamic Chart, 7=Column Dynamic Stack,
                # 8=Card Chart
                cht_info = {}
                chart_row_id_value = []
                chart_list = []
                data = []
                filtered_query = filter_conditions(
                    request, i.chart_query.get('sql_query'), i.filter_info, selected_items)
                chart_data = list(set_card_chart_data(
                    filtered_query, i.chart_query.get('labels')))
                cht_info = {"chart_type": "CARDCHART"}
                cht_info["chart_title"] = i.chart_title
                # cht_info.update({"colours": [
                                # {'role': 'style'}, '#FF3333', '#32B517', '#FFEA00', '#FFAA00', '#FF3333']})
                # chart_data = [('Male',96), ('Female',45), ('TG',32)]
                cht_info["datas"] = chart_data
                cht_info["bx_bg"] = i.chart_options.get('bx_bg',[])
                cht_info["bx_icon"] = i.chart_options.get('bx_icon',[])
                cht_info["bx_text"] = i.chart_options.get('bx_text',[])
                cht_info["bx_div"] = i.chart_options.get('bx_div',[])
                cht_info.update({"tooltip": i.chart_tooltip})
                cht_info.update({"chart_note": i.chart_note})
                cht_info.update({"chart_name": i.chart_slug})
                cht_info["chart_height"] = i.chart_height
                cht_info.update({"div": i.div_class})
                cht_info.update({"click_url_template":i.chart_query.get("click_url","")})
                chart_list.append(cht_info)
                counter_list = [i[1] for i in chart_data]
            elif i.chart_type == 10:#line chart
                cht_info = {}
                filtered_query = filter_conditions(
                    request, i.chart_query.get('sql_query'), i.filter_info, selected_items)
                
                chart_data = set_dynamic_table_chart_data(
                    filtered_query, headers)
                cht_info = {"chart_type": "LINECHART"}
                cht_info["chart_title"] = i.chart_title
                # cht_info.update({"colours": [
                                # {'role': 'style'}, '#FF3333', '#32B517', '#FFEA00', '#FFAA00', '#FF3333']})
                # chart_data = [('Male',96), ('Female',45), ('TG',32)]
                cht_info["datas"] = chart_data
                cht_info["bx_bg"] = i.chart_options.get('bx_bg',[])
                cht_info["bx_icon"] = i.chart_options.get('bx_icon',[])
                cht_info["bx_text"] = i.chart_options.get('bx_text',[])
                cht_info["bx_div"] = i.chart_options.get('bx_div',[])
                cht_info.update({"tooltip": i.chart_tooltip})
                cht_info.update({"chart_note": i.chart_note})
                cht_info.update({"chart_name": i.chart_slug})
                cht_info["chart_height"] = i.chart_height
                cht_info.update({"div": i.div_class})
                cht_info.update({"click_url_template":i.chart_query.get("click_url","")})
                chart_list.append(cht_info)
            elif i.chart_type == 1:
                # chart_type values: 
                # 1=Column Chart, 2=Pie Chart, 3=Table Chart, 4=Bar Chart, 5=Column Stack, 6=Bar Dynamic Chart, 7=Column Dynamic Stack,
                cht_info = {}
                # list to hold chart row id value - location id or classification id, etc based for the row
                # row/chart number mapps to the index in the list
                # used for dynamic charts where click handling is required
                chart_row_id_value = []
                filtered_query = filter_conditions(
                    request, i.chart_query.get('sql_query'), i. filter_info, selected_items)
                chart_data = list(set_column_chart_data(
                    filtered_query, i.chart_query.get('labels')))
                cht_info = {"chart_type": "COLUMNCHART"}
                cht_info["chart_title"] = i.chart_title
                chart_data.insert(0, ('', ''))
                cht_info["datas"] = chart_data
                # cht_info["datas"] = [('', ''), ('Ordered', 60), ('Pending', 30), ('Ready', 40), ('Delivered', 50), (' Received', 25),('Examined',55),('Prescibed',10)] 
                cht_info.update({"options": i.chart_options})
                cht_info.update({"colours": [
                                {'role': 'style'}, '#FF3333', '#32B517', '#FFEA00', '#FFAA00', '#FF3333']})
                cht_info.update({"tooltip": i.chart_tooltip})
                cht_info.update({"chart_note": i.chart_note})
                cht_info.update({"chart_name": i.chart_slug})
                cht_info["chart_height"] = i.chart_height
                cht_info.update({"div": i.div_class})
                cht_info.update({"click_url_template":i.chart_query.get("click_url","")})
                cht_info.update({"chart_row_id_value":chart_row_id_value})
                chart_list.append(cht_info)
            elif i.chart_type == 2:
                # chart_type values: 
                # 1=Column Chart, 2=Pie Chart, 3=Table Chart, 4=Bar Chart, 5=Column Stack, 6=Bar Dynamic Chart, 7=Column Dynamic Stack,
                cht_info = {}
                # list to hold chart row id value - location id or classification id, etc based for the row
                # row/chart number mapps to the index in the list
                # used for dynamic charts where click handling is required
                chart_row_id_value = []
                data = []
                filtered_query = filter_conditions(
                    request, i.chart_query.get('sql_query'), i. filter_info, selected_items)

                chart_data = list(set_pie_chart_data(
                    filtered_query, i.chart_query.get('labels')))

                cht_info = {"chart_type": "PIECHART"}
                cht_info["chart_title"] = i.chart_title
                cht_info.update({"colours": [
                                {'role': 'style'}, '#FF3333', '#32B517', '#FFEA00', '#FFAA00', '#FF3333']})
                chart_data.insert(0, ('', ''))
                cht_info["datas"] = chart_data
                # cht_info["datas"] = [('', ''), ('Male', 60), ('Female', 50), ('Transgender', 40)]
                cht_info["options"] = i.chart_options
                cht_info["options"].update({"sliceVisibilityThreshold":0.0001})
                cht_info.update({"tooltip": i.chart_tooltip})
                cht_info.update({"chart_note": i.chart_note})
                cht_info.update({"chart_name": i.chart_slug})
                cht_info["chart_height"] = i.chart_height
                cht_info.update({"div": i.div_class})
                cht_info.update({"click_url_template":i.chart_query.get("click_url","")})
                chart_list.append(cht_info)
            elif i.chart_type == 3:
                # chart_type values: 
                # 1=Column Chart, 2=Pie Chart, 3=Table Chart, 4=Bar Chart, 5=Column Stack, 6=Bar Dynamic Chart, 7=Column Dynamic Stack,
                cht_info = {"chart_type": "TABLECHART"}
                # list to hold chart row id value - location id or classification id, etc based for the row
                # row/chart number mapps to the index in the list
                # used for dynamic charts where click handling is required
                # chart_row_id_value = []
                headers = i.chart_query.get('col_headers')
                filtered_query = filter_conditions(
                    request, i.chart_query.get('sql_query'), i. filter_info, selected_items)
                chart_data = set_dynamic_table_chart_data(
                    filtered_query, headers)
                cht_info["chart_title"] = i.chart_title
                cht_info["options"] = i.chart_options
                cht_info["datas"] = chart_data
                # cht_info["datas"] = [('Visual impairment', 'Total No. of <br/>Patients'), ('Early VI (6/12-6/18) (Better eye PV)', 4), ('Moderate VI(6/18-6/60) (Better eye PV)', 6), ('Severe VI(6/60-3/60) (Better eye PV)', 5), ('Blind (less than 3/60) (Better eye PV)', 0)]
                # cht_info["datas"] = [('Spectacle Type', 'gender','Ordered','Pending','Ready','Delivered','Received'), ('Near','Male', 60, 12, 18, 45, 19), ('R2C','Female', 50,80,22,44,10), ('R2C','Transgender', 40,44,19,25,16), ('Near','Male', 60, 12, 18, 45, 19), ('R2C','Male', 60, 12, 18, 45, 19),('R2C','Female', 60, 12, 19, 45, 52),('R2C','Male', 45, 12, 18, 45, 19)]
                cht_info["tooltip"] = i.chart_tooltip
                cht_info["chart_height"] = i.chart_height
                cht_info["chart_name"] = i.chart_slug
                cht_info["div"] = i.div_class
                cht_info.update({"click_url_template":i.chart_query.get("click_url","")})
                # cht_info.update({"chart_row_id_value":chart_row_id_value})
            elif i.chart_type == 4 or i.chart_type == 6:
                # chart_type values: 
                # 1=Column Chart, 2=Pie Chart, 3=Table Chart, 4=Bar Chart, 5=Column Stack, 6=Bar Dynamic Chart, 7=Column Dynamic Stack,
                cht_info = {}
                # list to hold chart row id value - location id or classification id, etc based for the row
                # row/chart number mapps to the index in the list
                # used for dynamic charts where click handling is required
                chart_row_id_value = []
                filtered_query = filter_conditions(
                    request, i.chart_query.get('sql_query'), i. filter_info, selected_items)
                if i.chart_type == 4:
                    chart_data = list(set_pie_chart_data(
                        filtered_query, i.chart_query.get('labels')))
                else:
                    # chart_data, chart_row_id_value = list(set_bar_chart_dynamic_lable(filtered_query))
                    first_tuple, *rest_of_data = [('Place', 'Value'), ('Vision Center', 180), ('Camp', 220), ('Both', 60)]
                cht_info = {"chart_type": "BARCHART"}
                cht_info["chart_title"] = i.chart_title
                chart_data.insert(0, ('', ''))
                
                cht_info["datas"] = chart_data
                # cht_info["datas"] = [('Place', 'Value'), ('Vision Center', 180), ('Camp', 220), ('Both', 60)]
                cht_info.update({"tooltip": i.chart_tooltip})
                cht_info.update({"options": i.chart_options})
                cht_info.update({"chart_note": i.chart_note})
                cht_info.update({"chart_name": i.chart_slug})
                cht_info["chart_height"] = i.chart_height
                cht_info.update({"div": i.div_class})
                cht_info.update({"click_url_template":i.chart_query.get("click_url","")})
                #location id values. chart bar mapped to the index in list
                cht_info.update({"chart_row_id_value":chart_row_id_value})
                #logger.error("URL_Template:"+ i.chart_query.get("click_url",""))
                chart_list.append(cht_info)
            elif i.chart_type == 5:
                # chart_type values: 
                # 1=Column Chart, 2=Pie Chart, 3=Table Chart, 4=Bar Chart, 5=Column Stack, 6=Bar Dynamic Chart, 7=Column Dynamic Stack,
                cht_info = {"chart_type": "COLUMNSTACK"}
                # list to hold chart row id value - location id or classification id, etc based for the row
                # row/chart number mapps to the index in the list
                # used for dynamic charts where click handling is required
                chart_row_id_value = []
                filtered_query = filter_conditions(
                    request, i.chart_query.get('sql_query'), i.filter_info, selected_items)
                headers = i.chart_query.get('col_headers')
                chart_data = set_dynamic_table_chart_data(
                    filtered_query, headers)
                cht_info["chart_title"] = i.chart_title
                cht_info["datas"] = chart_data
                cht_info["options"] = i.chart_options
                # chart_type values: 1=Column Chart, 2=Pie Chart, 3=Table Chart , 4- Column Stack
                cht_info["chart_height"] = i.chart_height
                cht_info["chart_name"] = i.chart_slug
                cht_info["div"] = i.div_class
                cht_info.update({"tooltip": i.chart_tooltip})
                cht_info.update({"chart_note": i.chart_note})
                cht_info.update({"click_url_template":i.chart_query.get("click_url","")})
                cht_info.update({"chart_row_id_value":chart_row_id_value})
                chart_list.append(cht_info)
            elif i.chart_type == 9:
                # chart_type values: 
                # 1=Column Chart, 2=Pie Chart, 3=Table Chart, 4=Bar Chart, 5=Column Stack, 6=Bar Dynamic Chart, 7=Column Dynamic Stack, 8=Card Chart, 9=HTML Table Chart,10=Line chart,11=Progressive Line, 12=Dounut chart
                cht_info = {"chart_type": "HTMLTABLECHART"}
                cht_info["headers"] = i.chart_query.get('thead')
                cht_info["chart_title"] = i.chart_title
                filtered_query = filter_conditions(
                    request, i.chart_query.get('sql_query'), i.filter_info, selected_items)
                logger.error("HTMLTABLECHART QUERY: " + filtered_query)
                url_columns =  i.chart_query.get('url_columns',[])
                cht_info["url_columns"] = url_columns
                chart_data = table_html_chart(filtered_query)
                cht_info["datas"] = chart_data
                # cht_info["options"] = i.chart_options
                # # chart_type values: 1=Column Chart, 2=Pie Chart, 3=Table Chart , 4- Column Stack
                cht_info["chart_height"] = i.chart_height
                cht_info["chart_name"] = i.chart_slug
                cht_info["div"] = i.div_class
                cht_info.update({"tooltip": i.chart_tooltip})
                cht_info.update({"chart_note": i.chart_note})
                cht_info.update({"row_css_class":i.chart_query.get("row_css_class",{})})
                color_code_columns = i.chart_query.get("color_coded_columns",{})
                num_header_rows_in_data_list = color_code_columns.get("num_header_rows_in_data_list",0)
                color_codes = {}
                if color_code_columns != {}:
                    # logger.error("color_code_columns:"+str(color_code_columns))
                    # logger.error("color_code_columns-type:"+str(type(color_code_columns)))
                    color_codes = generate_color_codes(color_code_columns, chart_data, num_header_rows_in_data_list)
                    # logger.error("color_codes:" + str(color_codes))
                cht_info.update({"color_codes":color_codes})
                # cht_info.update({"click_url_template":i.chart_query.get("click_url","")})
                # cht_info.update({"chart_row_id_value":chart_row_id_value})
                chart_list.append(cht_info)
            elif i.chart_type == 14:
                # chart_type values: 
                # 1=Column Chart, 2=Pie Chart, 3=Table Chart, 4=Bar Chart, 5=Column Stack, 6=Bar Dynamic Chart, 7=Column Dynamic Stack, 8=Card Chart, 9=HTML Table Chart,10=Line chart,11=Progressive Line, 12=Dounut chart
                cht_info = {"chart_type": "CUSTOMHTMLTABLECHART"}
                header_query = i.chart_query.get('header_dynamic')
                cht_info["chart_title"] = i.chart_title
                report_headers_filters = getfiniacialmonth(selected_items[0],selected_items[1])
                sql_query = i.chart_query.get('sql_query').replace('@@start_date',f"'{report_headers_filters[1]}'").replace('@@end_date',f"'{report_headers_filters[2]}'")
                filtered_query = filter_conditions(
                    request, sql_query, i.filter_info, selected_items)
                print(filtered_query,'----------')
                logger.error("CUSTOMHTMLTABLECHART QUERY: " + filtered_query)
                # url_columns =  i.chart_query.get('url_columns',[])
                # cht_info["url_columns"] = url_columns
                chart_data = custom_table_html_chart(filtered_query)
                chart_header_data = custom_header_html_chart(header_query)
                print(selected_items[0],selected_items[1],'-----------------')
                # chart_data_with_url = generate_url_pattern(request,filtered_query,selected_items,url_columns,chart_data,i.chart_slug)
                cht_info["datas"] = chart_data
                cht_info["headers"] = report_headers_filters[0]
                # cht_info["options"] = i.chart_options
                # # chart_type values: 1=Column Chart, 2=Pie Chart, 3=Table Chart , 4- Column Stack
                cht_info["chart_height"] = i.chart_height
                cht_info["chart_name"] = i.chart_slug
                cht_info["div"] = i.div_class
                cht_info.update({"tooltip": i.chart_tooltip})
                cht_info.update({"chart_note": i.chart_note})
                cht_info.update({"row_css_class":i.chart_query.get("row_css_class",{})})
                color_code_columns = i.chart_query.get("color_coded_columns",{})
                num_header_rows_in_data_list = color_code_columns.get("num_header_rows_in_data_list",0)
                color_codes = {}
                if color_code_columns != {}:
                    # logger.error("color_code_columns:"+str(color_code_columns))
                    # logger.error("color_code_columns-type:"+str(type(color_code_columns)))
                    color_codes = generate_color_codes(color_code_columns, chart_data, num_header_rows_in_data_list)
                    # logger.error("color_codes:" + str(color_codes))
                cht_info.update({"color_codes":color_codes})
                # cht_info.update({"click_url_template":i.chart_query.get("click_url","")})
                # cht_info.update({"chart_row_id_value":chart_row_id_value})
                chart_list.append(cht_info)
            elif i.chart_type == 13:
                # chart_type values: 
                #1=Column Chart, 2=Pie Chart, 3=Table Chart, 4=Bar Chart, 5=Column Stack, 6=Bar Dynamic Chart, 7=Column Dynamic Stack, 8=Card Chart, 9=Geo chart,10=Line chart,11=Progressive Line, 12=Dounut chart,13= 'Html Table Chart Dynmaic'
                cht_info = {"chart_type": "HTMLTABLECHARTDYNAMIC"}
                header_query = i.chart_query.get('header_dynamic')
                header_query = replace_filter_values(header_query, selected_items)
                header_query = filter_conditions(request, header_query, i.filter_info, selected_items)
                dynamic_headers = custom_header_query(header_query)
                cht_info["headers"] = dynamic_headers
                cht_info["chart_title"] = i.chart_title
                filtered_query = i.chart_query.get('sql_query')
                filtered_query = filter_conditions(
                    request, i.chart_query.get('sql_query'), i.filter_info, selected_items)
                filtered_query = replace_filter_values(filtered_query,selected_items)
                logger.error("HTMLTABLECHARTDYNAMIC QUERY: " + filtered_query)
                chart_data = dynamic_table_html_chart(filtered_query)
                chart_data_temp = []
                if chart_data[0][0] is not None:
                    chart_data_temp = [[ch_data['row_headers']] + ch_data['row_data'] for ch_data in json.loads(chart_data[0][0])]
                # chart_data_temp = [[]]
                cht_info["datas"] = chart_data_temp
                # cht_info["options"] = i.chart_options
                # # chart_type values: 1=Column Chart, 2=Pie Chart, 3=Table Chart , 4- Column Stack
                cht_info["chart_height"] = i.chart_height
                cht_info["chart_name"] = i.chart_slug
                cht_info["div"] = i.div_class
                cht_info.update({"tooltip": i.chart_tooltip})
                cht_info.update({"chart_note": i.chart_note})
                url_columns = i.chart_query.get("url_columns",[])
                cht_info.update({"url_columns":url_columns})
                cht_info.update({"row_css_class":i.chart_query.get("row_css_class",{})})
                color_code_columns = i.chart_query.get("color_coded_columns",{})
                num_header_rows_in_data_list = color_code_columns.get("num_header_rows_in_data_list",0)
                color_codes = {}
                if color_code_columns != {} and chart_data_temp != []:
                    #logger.error("color_code_columns:"+str(color_code_columns))
                    #logger.error("color_code_columns-type:"+str(type(color_code_columns)))
                    color_codes = generate_color_codes(color_code_columns, chart_data_temp, num_header_rows_in_data_list)
                    #logger.error("color_codes:" + str(color_codes))
                cht_info.update({"color_codes":color_codes})
                # cht_info.update({"click_url_template":i.chart_query.get("click_url","")})
                # cht_info.update({"chart_row_id_value":chart_row_id_value})
                chart_list.append(cht_info)
                #logger.error("ChartData:" + str(chart_data_temp))
            #logger.error("SLUG: " + str(i.chart_slug) + " -- filtered_query" + str(filtered_query))
        data = {"chart": chart_list}
        #store the dashboard filter values for handling chart click events
        
        # mat_view_last_updated = DashboardWidgetSummaryLog.objects.get(
        #     status=2, log_key='meta_dashboard_views').last_successful_update
        request_data = None
        context = {
            'data': json.dumps(data, cls=DateTimeEncoder),
            'data_html': data,
            'page_slug': page_slug,
            'selected_items': selected_items,
            'partnerlist':partnerlist,
            # 'levels_to_filter':levels_to_filter,
            'request':request,
            'selected_partner_list':selected_partner_list,
            'selected_district_list':selected_district_list,
            'districts_list':district_list,
            'partner_list_id':partner_list_id,
            # 'districtlist':districtlist,
            'district_list_id':district_list_id,
            'financial_year_list':financial_year_list,
            'quarterly_year_list':quarterly_year_list
        }
        # if page_slug == 'mis-dashboard1':
        #     return render(request, 'dashboard/dashboard.html', context)
        # elif page_slug == 'block-wise-summary':
        #     return render(request, 'dashboard/dashboard_block_wise_head_wise.html', context)
        # else:
        #     # return export_excel(request,chart_list)
        if page_slug == 'mhm-dashboard':
            return render(request, 'dashboard/dashboard_mhm.html', context)
        else:
            return render(request, 'dashboard/dashboard.html', context)
    except KeyError:
        return redirect('/login/')


def getfiniacialmonth(financial_year,q_year):
    months = []
    months_header = []
    month_start = f"Apr-{financial_year.split('-')[0][2:]}"
    start_month = datetime.strptime(month_start, "%b-%y")
    if q_year == 'Q1':
        length=3
        for i in range(0,length):
            next_month = start_month + relativedelta(months=i)
            months.insert(i,next_month.strftime("%Y-%m-%d"))
            months_header.insert(i,next_month.strftime("%b-%y"))
        months_header.insert(i+1,'Quarter 1')
        months_header.insert(i+2,'Total')
        months_header.insert(i+3,'Cummulative Acheivemnet')
        result = (months_header,months[0],months[-1])
    elif q_year == 'Q2':
        length=6
        for i in range(0,length):
            next_month = start_month + relativedelta(months=i)
            months.insert(i,next_month.strftime("%Y-%m-%d"))
            months_header.insert(i,next_month.strftime("%b-%y"))
        months_header.insert(3,'Quarter 1')
        months_header.append('Quarter 2')
        months_header.append('Total')
        months_header.append('Cummulative Acheivemnet')
        result = (months_header,months[0],months[-1])
    elif q_year == 'Q3':
        length=9
        for i in range(0,length):
            next_month = start_month + relativedelta(months=i)
            months.insert(i,next_month.strftime("%Y-%m-%d"))
            months_header.insert(i,next_month.strftime("%b-%y"))
        months_header.insert(3,'Quarter 1')
        months_header.insert(7,'Quarter 2')
        months_header.append('Quarter 3')
        months_header.append('Total')
        months_header.append('Cummulative Acheivemnet')
        result = (months_header,months[0],months[-1])
    elif q_year == 'Q4':
        length=12
        for i in range(0,length):
            next_month = start_month + relativedelta(months=i)
            months.insert(i,next_month.strftime("%Y-%m-%d"))
            months_header.insert(i,next_month.strftime("%b-%y"))
        months_header.insert(3,'Quarter 1')
        months_header.insert(7,'Quarter 2')
        months_header.insert(11,'Quarter 3')
        months_header.append('Quarter 4')
        months_header.append('Total')
        months_header.append('Cummulative Acheivemnet')
        result = (months_header,months[0],months[-1])
    return result

def export_excel(request,chart_data):
    excel_meta_data = []
    excel_config = []
    for c_meta in chart_data:
        excel_config_dict = {}
        excel_data = []
        if len(c_meta['headers']) == 2:
            headers_1 = [head['label'] for head in c_meta['headers'][0]]
            headers_2 = [head['label'] for head in c_meta['headers'][1]]
            excel_data.insert(0,headers_1)
            excel_data.insert(1,headers_2)
        else:
            headers_1 = [head['label'] for head in c_meta['headers'][0]]
            excel_data.insert(0,headers_1)
        for idx,data in enumerate(c_meta['datas']):
            excel_data.insert(idx+2,list(data))
        excel_meta_data.append(excel_data)
        excel_config_dict[c_meta['chart_title']] = [len(excel_data),len(excel_data[-1])]
        excel_config.append(excel_config_dict)
    response = write_excel_data(request,excel_config,excel_meta_data)
    return response

def write_excel_data(request,excel_config,excel_meta_data):
    import openpyxl,os
    chart_position = {0:False,1:False,2:True,3:False}
    from django.http import HttpResponse
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Dashboard"
    dashboard_sheet = workbook.worksheets[0]
    for idx,row_col in enumerate(excel_config):
        if chart_position[idx] == True:
            row_idx = row_idx+value[0]+5
            col_idx = 3
            dashboard_sheet.cell(row=row_idx, column=col_idx, value='')
        elif idx != 0:
            row_idx = row_idx
            col_idx = col_idx+value[1]+5
            dashboard_sheet.cell(row=row_idx, column=col_idx, value='')
        else:
            row_idx = 2
            col_idx = 2
            dashboard_sheet.cell(row=row_idx, column=col_idx, value='')
        for key,value in row_col.items():
            for data in excel_meta_data[idx]:
                dashboard_sheet.append(data)
    folder_file_name = "dashboard-mis"+"_"+ datetime.today().strftime("%d%m%y%H%M") + ".xlsx"
    file_path = os.path.join(settings.MEDIA_DIR + '/temp_dash_data/', folder_file_name)
    workbook.save(file_path)
    if os.path.exists(file_path):
        with open(file_path, "rb") as excel:
            data = excel.read()
            response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={folder_file_name}'
            os.remove(file_path)
    return response
def generate_color_codes(color_coded_columns, data_list, skip_num_rows):
    # column list, start and end columns are with index starting from 1
    column_indexes = []
    col_numbers_list = color_coded_columns.get("col_index_list",[])
    if col_numbers_list != []:
        column_indexes = col_numbers_list
    else:
        start_column_index = color_coded_columns.get("start_index",-1)
        end_column_index = color_coded_columns.get("end_index",-1)
        for i in range(1,len(data_list[0])+1):
            if i >= start_column_index and (end_column_index == -1 or i <= end_column_index):
                column_indexes.append(i)
    percent_values = []
    original_p_values= []
    p_val_for_max = []
    for idx,row in enumerate(data_list):
        if idx >= skip_num_rows: 
            for col_index in column_indexes:
                original_p_values.append(row[col_index-1])            
                p_val = round(int(row[col_index-1].strip('%'))*1.0/100.0,2) if row[col_index-1] is not None else row[col_index-1]            
                if p_val is not None and p_val != 0:
                    percent_values.append((p_val,row[col_index-1]))
                    p_val_for_max.append(p_val)
    percent_values = list(set(percent_values)) # to remove duplicates
    # logger.error("percent_values:" + str(percent_values))
    # logger.error("original_p_values:" + str(original_p_values))
    color_codes = {}
    if p_val_for_max and len(p_val_for_max) > 0:
        max_percentage_val = max(p_val_for_max)
        # Assign colors to percentages
        for p_t in percent_values:
            #color_codes.update({str(p).replace(".","").lstrip('0') + "%":percentage_to_color(p,max_percentage_val)})
            color_codes.update({p_t[1]:percentage_to_color(p_t[0],max_percentage_val)})
    return color_codes

# def percentage_to_color(percentage):
#     """Map a percentage to a color from red to dark green."""
#     # Map the percentage to a color from red to dark green
#     segment = int(percentage * 5)
#     segment_remainder = (percentage * 5) - segment
#     # Define colors for each segment
#     colors = [
#         (1.0, 0.0, 0.0),   # Red
#         (1.0, 0.5, 0.0),   # Orange
#         (1.0, 1.0, 0.0),   # Yellow
#         (0.5, 1.0, 0.0),   # Light Green
#         (0.0, 0.5, 0.0)    # Dark Green
#     ]
#     # Interpolate between adjacent colors
#     color = (
#         colors[segment][0] * (1 - segment_remainder) + colors[segment + 1][0] * segment_remainder,
#         colors[segment][1] * (1 - segment_remainder) + colors[segment + 1][1] * segment_remainder,
#         colors[segment][2] * (1 - segment_remainder) + colors[segment + 1][2] * segment_remainder
#     )
#     # Convert RGB to hexadecimal
#     color_code = "#{:02X}{:02X}{:02X}".format(int(color[0] * 255), int(color[1] * 255), int(color[2] * 255))
#     return color_code

def percentage_to_color(percentage, max_percentage_value):
    """Map a percentage to a color from red to dark green."""
    # Adjust the percentage to be within the range [0, 2], with .5 as yellow
    percentage = max(0.0, min(percentage, max_percentage_value))
    
    # Map the percentage to a color from red to dark green
    segment = int(percentage * 5 / 2)
    
    if segment >= 4:  # If the segment reaches the maximum
        #segment = 3
        segment_remainder = 1.0 if segment == 4 else 0.5
    else:
        segment_remainder = (percentage * 5 / 2) - segment

    # Define colors for each segment
    colors = [
        # (1.0, 0.0, 0.0),   # Red
        # (1.0, 0.5, 0.0),   # Orange
        # (1.0, 1.0, 0.0),   # Yellow
        # (0.5, 1.0, 0.0),   # Light Green
        # (0.0, 0.5, 0.0)    # Dark Green
        (1.0, 0.36, 0.37),   # Red 255,91,94
        (1.0, 0.66, 0.42),   # Orange 255,169,108
        (1.0, 0.94, 0.45),   # Yellow 255,239,114
        (0.65, 0.86, 0.46),   # Light Green 166,219, 118
        (0.22, 0.77, 0.46)    # Dark Green 55,197,119
    ]

    # Interpolate between adjacent colors
    if segment >= 4:  # If we're at the last segment
        color = colors[4]
    else:
        next_segment = segment+1 if segment < 4 else segment
        color = (
            colors[segment][0] * (1 - segment_remainder) + colors[next_segment][0] * segment_remainder,
            colors[segment][1] * (1 - segment_remainder) + colors[next_segment][1] * segment_remainder,
            colors[segment][2] * (1 - segment_remainder) + colors[next_segment][2] * segment_remainder
        )

    # Convert RGB to hexadecimal
    color_code = "#{:02X}{:02X}{:02X}".format(int(color[0] * 255), int(color[1] * 255), int(color[2] * 255))
    return color_code

class DateTimeEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, datetime):
            return o.isoformat()
        return super().default(o)

# ****************************************************************************
# Function to refresh_materialized_view
# ****************************************************************************

def refresh_materialized_view(view_name):
    query = f'REFRESH MATERIALIZED VIEW {view_name};'
    cursor = connection.cursor()
    cursor.execute(query)

def materialized_view_master():
    try:
        refresh_materialized_view('dash_glass_prescription_view')
        refresh_materialized_view('dash_patient_basic_info_view')
        refresh_materialized_view('dash_screening_info_view')
        print('MATERIALIZED VIEWS REFRESHED SUCCESS')
        now = datetime.now()
        logdata, created = DashboardWidgetSummaryLog.objects.get_or_create(
            log_key='meta_dashboard_views')
        logdata.last_successful_update = now
        logdata.most_recent_update = now
        logdata.most_recent_update_status = 'Success'
        logdata.save()
    except Exception as ex1:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        error_stack = repr(traceback.format_exception(
            exc_type, exc_value, exc_traceback))
        logger.error(error_stack)
        now = datetime.now()
        logdata, created = DashboardWidgetSummaryLog.objects.get_or_create(
            log_key='meta_dashboard_views')
        logdata.last_successful_update = now
        logdata.most_recent_update = now
        logdata.most_recent_update_status = 'Failed'
        logdata.save()
    return True

@login_required(login_url="/login/")
def monthly_dashboard_list(request):
    month = request.GET.get('month')
    submitted_on = request.GET.get('submitted_on')
    approval_status_id = request.GET.get('approval_status')
    approval_status = STATUS_CHOICES

    # user_partner = UserProjectMapping.objects.filter(active=2,user=request.user,project__application_type_id = 511).values_list('project__partner_mission_mapping__partner_id', flat=True)
    user_partner = request.session['user_partner_list']

    object_list = MonthlyDashboard.objects.filter(active=2,partner__in=user_partner).order_by('-modified').select_related('partner','project_incharge','partner_admin','submitted_by')
    if month:
        month_int = ''.join(month.split('-')[::-1])
        object_list = object_list.filter(month = month_int)
    
    if submitted_on:
        object_list = object_list.filter(created__date = submitted_on)
    
    if approval_status_id:
        object_list = object_list.filter(current_status = approval_status_id)

    object_list = get_pagination(request, object_list)
    return render(request, "survey_forms/activity_list.html", locals())



@csrf_exempt
@login_required(login_url="/login/")
def dashboard_data_approval(request, id):
    from survey.api_views_version1 import submitted_record_mails
    try:
        monthly_data = MonthlyDashboard.objects.get(id=id)
        month_obj = datetime.strptime(str(monthly_data.month), 'Y%m%d')
    except:
        current_date = datetime.now()
        month_obj,end_of_previous_month=get_first_and_last_date_of_month(current_date.year,DASHBOARD_SUBMISSION_DAY)
        # first_day_of_current_month = current_date.replace(day=1)
        # first_day_of_current_month = current_date.replace(month=DASHBOARD_SUBMISSION_DAY)
        # month_obj = first_day_of_current_month - timedelta(days=1)
        # month_obj = first_day_of_current_month.replace(day=1)
        # start_of_previous_month = month_obj.replace(day=1)
        # print(start_of_previous_month.date(),month_obj.date())
        monthly_data_queryset = JsonAnswer.objects.values(
            'cluster__BeneficiaryResponse'
        ).filter(
            active=2,
            survey_id__in=[6, 8, 5, 4, 3, 10],
            created__date__range=[month_obj,end_of_previous_month]
        ).annotate(
            primary_screening_done=Sum(
                Case(
                    When(
                        (~Q(response__contains={"407": ""})),
                        survey_id=3,
                        response__has_key="407",
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            secondary_screening_done=Sum(
                Case(
                    When(
                        (~Q(response__contains={"405": ""})),
                        survey_id=4,
                        response__has_key="405",
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            teachers_training_done=Sum(
                Case(
                    When(
                        survey_id=10,
                        response__contains={"393": "262"},
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            teachers_training_notdone=Sum(
                Case(
                    When(
                        survey_id=10,
                        response__contains={"393": "263"},
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            rx_spec=Sum(
                Case(
                    When(
                        survey_id=4,
                        response__contains={"282": "176"},
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            prov_spec=Sum(
                Case(
                    When(
                        (~Q(response__contains={"355": ""})),
                        survey_id=8,
                        response__has_key="355",
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            advc_same_spec=Sum(
                Case(
                    When(
                        survey_id=4,
                        response__contains={"281": "173"},
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            ref_hosp_exam=Sum(
                Case(
                    When(
                        survey_id=4,
                        response__contains={"284": "180"},
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            prov_spec_hosp=Sum(
                Case(
                    When(
                        survey_id=4,
                        response__contains={"283": "179"},
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            advc_surgery=Sum(
                Case(
                    When(
                        survey_id=5,
                        response__contains={"343": "200"},
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            prov_surgery=Sum(
                Case(
                    When(
                        survey_id=6,
                        response__contains={"345": "204"},
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            swc_3months=Sum(
                Case(
                    When(
                        Q(response__contains={"356": "222"}) | Q(response__contains={"356": "224"}),
                        survey_id=8,
                        then=Value(1)
                    ),
                    default=Value(0),
                )
            ),
            combined_screening_done=Case(
                When(
                    Q(primary_screening_done__gte=1) & Q(secondary_screening_done__gte=1),
                    then=Value(1)
                ),
                default=Value(0),
            ),
            is_school=Case(
                When(survey__config__0__object_id_1='3', then=Value(1)),
                default=Value(0),
            ),
            is_student=Case(
                When(survey__config__0__object_id_1='2', then=Value(1)),
                default=Value(0),
            ),
            school_covered=Case(
                When(
                    Q(is_school=1) & Q(Q(combined_screening_done__gte=1) | Q(teachers_training_done__gte=1) | Q(teachers_training_notdone__gte=1)),
                    then=Value(1)
                ),
                default=Value(0),
            )
        )
        monthly_data_dict = monthly_data_queryset.aggregate(
            children_covered_count=Sum('combined_screening_done'),
            # sum_primary_screening_done=Sum('primary_screening_done'),
            # sum_secondary_screening_done=Sum('secondary_screening_done'),
            teachers_train_count=Sum('teachers_training_done'),
            # sum_teachers_training_notdone=Sum('teachers_training_notdone'),
            children_pres_count=Sum('rx_spec'),
            child_prov_spec_count=Sum('prov_spec'),
            pgp_count=Sum('advc_same_spec'),
            children_reffered_count=Sum('ref_hosp_exam'),
            child_prov_hos_count=Sum('prov_spec_hosp'),
            children_adv_count=Sum('advc_surgery'),
            children_prov_sgy_count=Sum('prov_surgery'),
            swc_count=Sum('swc_3months'),
            school_covered_count=Sum('school_covered')
        )

        # Convert the queryset to a list of SimpleNamespace objects
        monthly_data = SimpleNamespace(**monthly_data_dict)
        
    user_group = request.session['user_group_list']
    
    # if record status have the permission of group to action {status:role id (group)}
    role_with_permisson_map = {1:4,2:2,3:3}

    if request.method == "POST":
        # 4 = Partner Admin
        # 1 = Partner Data Entry Operator
        # 2 = Project In-charge
        # 3 = MFV Admin
        button = request.POST.get('label')
        approval = {'reject':-1,'approve':+1}
        partner_list = request.session.get('user_partner_list')
        if id == 'new':
            array_fields = [
                field.name for field in MonthlyDashboard._meta.get_fields()
                if isinstance(field, ArrayField)
            ]
            for field in array_fields:
                monthly_data_dict[field] = []

            month_field = month_obj.strftime('Y%m%d')

            # Preprocess the dictionary to replace None with 0
            preprocessed_dict = {k: (0 if v is None else v) for k, v in monthly_data_dict.items()}

            monthly_data, _ = MonthlyDashboard.objects.update_or_create(month = month_field,creation_key = datetime.now().strftime('%Y%m%d%H%M%S') + str(uuid4()),partner_id=partner_list[0],submitted_by = request.user,defaults=preprocessed_dict)
            
            # removing the session true for this month
            del request.session['show_submission_popup']

        monthly_data.current_status += approval.get(button)
        
        if 4 in user_group:
            # if only partner admin rejected means it will be 5 (rejected)
            monthly_data.current_status = 5 if button == 'reject' else monthly_data.current_status
            monthly_data.partner_admin = request.user
            monthly_data.partner_submitted = datetime.today()
        elif 2 in user_group:
            monthly_data.project_incharge = request.user
            monthly_data.project_incharge_submitted = datetime.today()

        monthly_data.save()
        if request.POST.get('remark'):
            group = request.user.groups.all()[0]
            remark_text = f"{request.user.username} ({group.name})  -  {request.POST.get('remark')}"
            Remarks.objects.create(object_id=monthly_data.id,content_type_id=62,remark=remark_text,user=request.user)
        return JsonResponse({'message': 'Updated successfully'})

    dashboard_data = {"No. of Children Screened":monthly_data.children_covered_count,"No. of Schools Covered":monthly_data.school_covered_count,"No. of Teachers Trained":monthly_data.teachers_train_count,"No. of Children Prescribed Spectacles":monthly_data.children_pres_count,"No. of Children Provided Spectacles":monthly_data.child_prov_spec_count,"No. of Children Advised to Continue with Same Glasses (PGP)":monthly_data.pgp_count,"No. of Children Referred to Hospital for Detailed Examination":monthly_data.children_reffered_count,"No. of Children Provided Spectacles at Hospital":monthly_data.child_prov_hos_count,"No. of Children Advised Surgery":monthly_data.children_adv_count,"No. of Children Provided Surgery":monthly_data.children_prov_sgy_count,"Spectacle Wearing Compliance After 3 Months":monthly_data.swc_count}

    remarks = Remarks.objects.filter(active=2,content_type_id=62,object_id=monthly_data.id).order_by('created')
    return render(request, "survey_forms/activity_submition_view.html", locals())
