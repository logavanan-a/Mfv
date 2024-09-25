import os
from django.conf import settings
import pandas as pd
import traceback
import sys
from django.core.exceptions import PermissionDenied
from django.shortcuts import render, redirect,get_list_or_404
from django.contrib.auth.decorators import login_required
from django.db import connection
from .models import ReportMeta,QuarterlyReport
from dashboard.models import DashboardSummaryLog
from django.contrib.contenttypes.models import ContentType
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from dateutil.relativedelta import relativedelta
import copy
import json
import re
import csv
import logging
from survey.models import BeneficiaryResponse
from django.contrib.auth.models import User
from datetime import datetime as dt
from dashboard.models import *
from application_master.models import (District, Donor, Menus, Mission,
                                       MissionIndicator,
                                       MissionIndicatorCategory,
                                       MissionIndicatorTarget, Partner,
                                       PartnerMissionMapping, Project,
                                       ProjectDonorMapping, ProjectFiles,
                                       State, UserPartnerMapping, UserProjectMapping)
from datetime import datetime
from openpyxl import Workbook
from openpyxl.writer.excel import save_workbook
from django.utils.encoding import escape_uri_path
import openpyxl

logger = logging.getLogger(__name__)
# ****************************************************************************
# execute Raw SQL
# ****************************************************************************


def return_sql_results(sql):
    """
    Executes an SQL query and returns the result rows.
    """
    cursor = connection.cursor()
    cursor.execute(sql)
    rows = cursor.fetchall()
    return rows

def return_sql_results_json(sql):
    """
    Executes an SQL query and returns the result rows.
    """
    result_list = []
    cursor = connection.cursor()
    cursor.execute(sql)
    rows = cursor.fetchall()
    # Assuming the query returns a single JSON array as a single row and column
    json_array_str = rows[0][0]
    if json_array_str:
    # Convert JSON string to a Python list
        result_list = json.loads(json_array_str)
    
    return result_list


# ****************************************************************************
# Excel function for export
# ****************************************************************************


def write_to_excel_from_normalized_table(conn_str, sql_query, headers_list, custom_export_header, rows_in_chunk, sheetname, excelWriter):
    # read from sql table into the pandas dataframe in chucks, this returns a list of dataframes - one for each chunk

    # create headers rows and write to excel sheet
    header_row_count = 1
    row_data = {}
    if custom_export_header:
        # custom export header is used to add custom header info - mostly multi level headers like the child classification report
        for i in custom_export_header:
            row_data.update({tuple(i): {}})
        header_row_count = len(custom_export_header[0])
        header_df = pd.DataFrame.from_dict(row_data)
    else:
        # when custom export header is None, then add the report headers as the export excel headers
        col_list = []
        for i in headers_list[0]:
            # remove any <br/> added in the header as formatting on the html pages
            header_text = i.get('label', '').replace('<br/>', '')
            headers_text = header_text.replace("%",'%%')
            col_list.append(header_text)
        header_df = pd.DataFrame([], columns=col_list)
        header_row_count = 1
        sql_query = sql_query.replace("%",'%%')
    normalized_df_list = pd.read_sql_query(
        sql_query, con=conn_str, chunksize=rows_in_chunk)
    start_row = header_row_count
    # set header to False indicating not to add the header data/row
    header_info = False
    df_count = 0
    try:
        for chunk_df in normalized_df_list:
            chunk_df.index += (df_count * rows_in_chunk)+1
            df_count = df_count + 1
            chunk_df.to_excel(excelWriter, sheet_name=sheetname,
                              startrow=start_row, header=header_info)
            # add the dataframe size (rows read from table/ chunk size) and header rows count (1 for first iteration and 0 for further iterations)
            start_row = start_row + chunk_df.shape[0] #+ header_row_count
            # set flags to indicate first dataframe is written to excel
            # set header row count used in the calcuation of start row to 0 as no further header rows will be added
            header_row_count = 0
            # unset flag for header info to false as no further header details will be written to sheet
            header_info = False

        # add the header row rowas at row 0 - insert it as a new data frame with empty data and the header rows
        header_df.to_excel(excelWriter, sheet_name=sheetname,
                           startrow=0)
    except Exception as e:
        df_count = 0
        exc_type, exc_value, exc_traceback = sys.exc_info()
        error_stack = repr(traceback.format_exception(
            exc_type, exc_value, exc_traceback))
        logger.error(error_stack)
        raise

@ login_required(login_url='/login/')
def reports_listing(request):
    """
    View function to render the reports listing page.
    """
    heading='Reports'
    page_reports = ReportMeta.objects.filter(active=2).order_by('display_order')
    return render(request, 'reports/reports.html', locals())
    

# @ login_required(login_url='/login/')
# def custom_report(request, page_slug):
#     rows_per_page = 10
#     if request.method == "POST":
#         req_data = request.POST
#     elif request.method == "GET":
#         req_data = request.GET
#     mat_view_last_updated = DashboardSummaryLog.objects.get(
#         active=2, log_key='mat_partner_mission_meta_view').last_successful_update
#     export_flag = True if req_data.get('export') and req_data.get(
#         'export').lower() == 'true' else False
#     # order of reports is specified in the ReportMeta default ordering
#     # page_reports = get_list_or_404(ReportMeta, page_slug=page_slug, active=2)
#     page_reports = ReportMeta.objects.filter(
#         page_slug=page_slug, active=2).order_by('display_order')
#     # temp veriable
#     data_query_list = []
#     #report_tabs = []
#     section_title = []
#     table_header = []
#     custom_export_headers = []
#     total_header_cols = []
#     report_slug_list = []
#     data = []
#     nowrap_cols = []
#     user_sort_field = []
#     user_sort_order = []
#     page_info = []
#     sorting_field = []
#     user_location_data = None
#     filter_values = None

#     # get user selected filter data and merge with the user configured location heirarcy
#     for idx, report in enumerate(page_reports):
#         r_slug = report.report_slug
#         s_title = report.report_title
#         f_info = report.filter_info
#         s_info = report.sort_info
#         r_query = report.report_query
#         d_query = r_query['sql_query']
#         c_query = r_query['count_query'] if r_query['count_query'] else ''
#         nw_cols = r_query['nowrap_cols']
#         mission_id = r_query['mission_id']
#         headers = report.report_header
#         e_header = report.custom_export_header
#         # all filter settings are set based on the first report filters
#         if idx == 0:
#             page_slug = report.page_slug
#             user_location_data, filter_values, user_filter_values, extended_filter_dict = get_filter_data(
#                 request, req_data, f_info,mission_id)
#         # update any variable_location_names  - in query, count query, sort and headers
#         default_sort = r_query['default_sort'] if 'default_sort' in r_query else None


#         headers, e_header, data_query, count_query, s_info, default_sort = apply_variable_location_info(
#             headers, e_header, d_query, c_query, s_info, default_sort, user_filter_values, user_location_data)

#         sort_field, sort_order = set_sort_options(
#             req_data, idx, s_info, default_sort)
#         user_sort_field.append(sort_field)
#         user_sort_order.append(sort_order)

#         # page reloads on click of filter, so current page is always set to 1
#         current_page = 1
#         data_query, count_query = apply_filters_to_query(
#             data_query, count_query, f_info, sort_field, sort_order, user_filter_values, current_page, rows_per_page, extended_filter_dict, export_flag)
#         # table header details
#         table_header.append(headers)
#         # get total columns count (colspan sum) - used to display the no records found row
#         header_col_count = 0
#         for item in headers[0]:
#             colspan = item.get('colspan', 0)
#             header_col_count += colspan if colspan > 0 else 1

#         data_query_list.append(data_query)
@ login_required(login_url='/login/')
def quarterly_report(request):
    heading = 'Quietly Report'
    from datetime import datetime
    current_year = datetime.now().year+1
    academic_year_list = [year for year in range(2022,current_year)]
    indicator_obj = MasterLookUp.objects.filter(parent_id=512,active=2)
    district_obj = District.objects.filter(active=2,id__in= Project.objects.filter(active=2).values_list('district_id',flat=True))
    district_id = request.GET.get('district','')
    project_id = None
    if district_id:
        projetc_obj = Project.objects.filter(active=2,district_id=district_id)
        project_id = projetc_obj.first().id if projetc_obj.exists() else ''
    academic_year = request.GET.get('academic_year','')
    if request.method == "POST":
        for indicator in indicator_obj:
            annual_target = request.POST.get('annual_target_'+str(indicator.id),0)
            q1_target = request.POST.get('q1_target_'+str(indicator.id),0)
            q2_target = request.POST.get('q2_target_'+str(indicator.id),0)
            q3_target = request.POST.get('q3_target_'+str(indicator.id),0)
            q4_target = request.POST.get('q4_target_'+str(indicator.id),0)
            obj, created=QuarterlyReport.objects.update_or_create(
                    project_id=project_id,
                    indicator_id=indicator.id,
                    academic_year=academic_year,
                    defaults={
                        "annual_target":annual_target or 0,
                        "q1_target":q1_target or 0,
                        "q2_target":q2_target  or 0,
                        "q3_target":q3_target  or 0,
                        "q4_target":q4_target or 0
                    }
                )
            obj.save()
        return redirect('/quarterly-report/?district='+str(district_id)+'&academic_year='+str(academic_year))

    return render(request, 'reports/quarterly_report.html', locals())
    
@ login_required(login_url='/login/')
def custom_report(request, page_slug):
    rows_per_page = 10
    if request.method == "POST":
        req_data = request.POST
    elif request.method == "GET":
        req_data = request.GET
    mat_view_last_updated = DashboardSummaryLog.objects.get(
        active=2, log_key='mat_partner_mission_meta_view').last_successful_update
    export_flag = True if req_data.get('export') and req_data.get(
        'export').lower() == 'true' else False
    # order of reports is specified in the ReportMeta default ordering
    # page_reports = get_list_or_404(ReportMeta, page_slug=page_slug, active=2)
    page_reports = ReportMeta.objects.filter(
        page_slug=page_slug, active=2).order_by('display_order')
    # temp veriable
    data_query_list = []
    #report_tabs = []
    section_title = []
    table_header = []
    custom_export_headers = []
    total_header_cols = []
    report_slug_list = []
    data = []
    nowrap_cols = []
    user_sort_field = []
    user_sort_order = []
    page_info = []
    sorting_field = []
    user_location_data = None
    filter_values = None
    # import ipdb;ipdb.set_trace()
    # get user selected filter data and merge with the user configured location heirarcy
    for idx, report in enumerate(page_reports):
        r_slug = report.report_slug
        s_title = report.report_title
        f_info = report.filter_info
        s_info = report.sort_info
        r_query = report.report_query
        d_query = r_query['sql_query']
        c_query = r_query['count_query'] if r_query['count_query'] else ''
        nw_cols = r_query['nowrap_cols']
        mission_id = r_query['mission_id']
        headers = report.report_header
        e_header = report.custom_export_header
        # all filter settings are set based on the first report filters
        if idx == 0:
            page_slug = report.page_slug
            user_location_data, filter_values, user_filter_values, extended_filter_dict = get_filter_data(
                request, req_data, f_info,mission_id)
        # update any variable_location_names  - in query, count query, sort and headers
        default_sort = r_query['default_sort'] if 'default_sort' in r_query else None


        headers, e_header, data_query, count_query, s_info, default_sort = apply_variable_location_info(
            headers, e_header, d_query, c_query, s_info, default_sort, user_filter_values, user_location_data)

        sort_field, sort_order = set_sort_options(
            req_data, idx, s_info, default_sort)
        user_sort_field.append(sort_field)
        user_sort_order.append(sort_order)

        # page reloads on click of filter, so current page is always set to 1
        current_page = 1
        data_query, count_query = apply_filters_to_query(
            data_query, count_query, f_info, sort_field, sort_order, user_filter_values, current_page, rows_per_page, extended_filter_dict, export_flag)
        # table header details
        table_header.append(headers)
        # get total columns count (colspan sum) - used to display the no records found row
        header_col_count = 0
        for item in headers[0]:
            colspan = item.get('colspan', 0)
            header_col_count += colspan if colspan > 0 else 1
        data_query_list.append(data_query)
        data.append(return_sql_results(data_query))
        total_header_cols.append(header_col_count)
        report_slug_list.append(r_slug)
        custom_export_headers.append(e_header)
        section_title.append(s_title)
        nowrap_cols.append(nw_cols)
        sorting_field.append(s_info)
        # if not for export, prepare pagination details
        if export_flag == False:
            # fetch record count and calcualte pagination info
            total_records = 0
            count_result=None
            # if request.method in ("POST","GET") and request.POST.get("filter"):
            count_result = return_sql_results(count_query)
            if count_result:
                total_records = count_result[0][0]
            p_info = calculate_pagination_info(total_records, 1, rows_per_page)
            page_info.append(p_info)
            # pagination_range will be same for all reports in the page
            pagination_range = range(p_info.get(
                'start_page'), p_info.get('end_page')+1)
    sidebar_active = 'Report'
    heading = section_title[0]
    # if export button click, create excel and return as response
    if export_flag == True:
        return generate_export_excel(section_title[0], data_query_list, table_header, custom_export_headers, section_title)
    return render(request, 'reports/multitab_report.html', locals())

def getacademicyear(fin_yaer):
    today = datetime.today()
    current_year = today.year
    diff_finyear = current_year - int(fin_yaer.split('-')[0])
    financial_year = []
    financial_year.append(fin_yaer)
    for year in range(diff_finyear):
        if int(financial_year[year].split('-')[1]) != current_year:
            financial_year.append(f"{financial_year[year].split('-')[1]}-{int(financial_year[year].split('-')[1])+1}")
        else:
            if today.month >= 4:
                financial_year.append(f"{financial_year[year].split('-')[1]}-{int(financial_year[year].split('-')[1])+1}")
    return financial_year

def getquarteryear(academic_year):
    selected_fy = academic_year
    selected_year = selected_fy.split('-')
    today = datetime.today()
    result_set = []
    current_year = today.year
    current_month = today.month
    if current_year == int(selected_year[0]) and current_month >= 7 and current_month <= 9:
        result_set.append('Q1')
    elif current_year == int(selected_year[0]) and current_month >= 10 and current_month <= 12:
        result_set.append('Q1')
        result_set.append('Q2')
        result_set.append('Q3')
    elif current_year == int(selected_year[1]) and current_month >= 1 and current_month <= 3:
        result_set.append('Q1')
        result_set.append('Q2')
        result_set.append('Q3')
    elif current_year == int(selected_year[1]) and current_month >= 4 and current_month <= 6:
        result_set.append('Q1')
        result_set.append('Q2')
        result_set.append('Q3')
        result_set.append('Q4')
    return result_set

def get_donor_district_list(request,partner_id,donor_id,district_id):
    user_id = request.user.id
    donor_list,district_list = [],[]
    user_details = load_user_details(request)
    if 'user_partner_list' in request.session:
        partner_ids = request.session['user_partner_list']
    partner_list = Partner.objects.filter(id__in=partner_ids).values_list('id','name').order_by('name')
    if partner_id != '' or len(partner_ids) == 1:
        if len(partner_ids) == 1: 
            part_mission_ids = PartnerMissionMapping.objects.filter(active=2,partner_id__in=partner_ids).values_list('id',flat=True)
        else:
            part_mission_ids = PartnerMissionMapping.objects.filter(active=2,partner_id=int(partner_id)).values_list('id',flat=True)
        project_ids = Project.objects.filter(active=2,partner_mission_mapping_id__in=part_mission_ids).values_list('id',flat=True)
        donor_ids = ProjectDonorMapping.objects.filter(active=2,project_id__in=project_ids).values_list('donor_id',flat=True)
        donor_list = Donor.objects.filter(active=2,id__in=donor_ids).values_list('id','name').order_by('name')
    if donor_id != '' and partner_id != '':
        part_mission_ids = PartnerMissionMapping.objects.filter(active=2,partner_id=int(partner_id)).values_list('id',flat=True)
        project_ids = Project.objects.filter(active=2,partner_mission_mapping_id__in=part_mission_ids).values_list('id',flat=True)
        donor_project_list = ProjectDonorMapping.objects.filter(active=2, project_id__in=project_ids).values_list('project_id', flat=True)
        district_project_list = Project.objects.filter(active=2, id__in=donor_project_list).values_list('district_id', flat=True)
        district_list_ids = District.objects.filter(active=2, id__in=district_project_list).values_list('id',flat=True)
        district_list_ids_str = list(map(str, district_list_ids))
        district_list = Boundary.objects.filter(active=2,code__in=district_list_ids_str,boundary_level_type_id=2).values_list('id','name').order_by('name')
    return partner_list,donor_list,district_list

@csrf_exempt
@ login_required(login_url='/login/')
def get_partner_district(request):
    if request.method == 'GET':
        selected_partner = request.GET.get('selected_partner', '')
        part_mission_ids = PartnerMissionMapping.objects.filter(active=2,mission_id = 2,partner_id=int(selected_partner)).values_list('id',flat=True)
        project_ids = Project.objects.filter(active=2,partner_mission_mapping_id__in=part_mission_ids).values_list('id',flat=True)
        donor_ids = ProjectDonorMapping.objects.filter(active=2,project_id__in=project_ids).values_list('donor_id',flat=True)
        donor_list = Donor.objects.filter(active=2,id__in=donor_ids).values_list('id','name').order_by('name')
        result_set = []
        for donor in donor_list:
            result_set.append(
                {'id': donor[0], 'name': donor[1], })
        return HttpResponse(json.dumps(result_set))



@csrf_exempt
@ login_required(login_url='/login/')
def get_donor_district(request):
    if request.method == 'GET':
        selected_donor = request.GET.get('selected_donor', '')
        selected_partner = request.GET.get('selected_partner','')
        part_mission_ids = PartnerMissionMapping.objects.filter(active=2,mission_id = 2,partner_id=int(selected_partner)).values_list('id',flat=True)
        #TODO added mission id 2 (Roshini) - 
        project_ids = Project.objects.filter(active=2,partner_mission_mapping_id__in=part_mission_ids).values_list('id',flat=True)
        donor_project_list = ProjectDonorMapping.objects.filter(active=2, project_id__in=project_ids,donor_id=int(selected_donor)).values_list('project_id', flat=True)
        district_project_list = Project.objects.filter(active=2, id__in=donor_project_list).values_list('district_id', flat=True)
        district_list_ids = District.objects.filter(active=2, id__in=district_project_list).values_list('id',flat=True)
        district_list_ids_str = list(map(str, district_list_ids))
        district_list = Boundary.objects.filter(active=2,code__in=district_list_ids_str,boundary_level_type_id=2).values_list('id','name').order_by('name')
        result_set = []
        for district in district_list:
            result_set.append(
                {'id': district[0], 'name': district[1], })
        return HttpResponse(json.dumps(result_set))

@ login_required(login_url='/login/')
def custom_report_donor(request):
    # import ipdb;ipdb.set_trace()
    # locations_data = load_user_details_to_session(request)
    academic_year_list = getacademicyear('2024-2025')
    current_academic_year = academic_year_list[-1]
    today = datetime.today()
    q_year = request.POST.get('q_year','q1')
    academic_year = request.POST.get('finance_year',current_academic_year)
    partner_id = request.POST.get('partner','')
    donor_id = request.POST.get('donor','')
    district_id = request.POST.get('district','')
    partner_list,donor_list,district_list = get_donor_district_list(request,partner_id,donor_id,district_id)
    financial_year = current_academic_year
    quarterly_year_list = getquarteryear(academic_year)
    no_table = True
    # state_cond,dist_cond,selected_state,selected_district,state_filter,district_filter = get_state_dist_cond(request)
    if request.method == 'POST':
        child_scr_data = child_screening_data(academic_year,q_year,partner_id,donor_id,district_id)
        spec_comp_data =spec_compliance_data(academic_year,q_year,partner_id,donor_id,district_id)
        teach_train_data =teacher_trained_data(academic_year,q_year,partner_id,donor_id,district_id)
        surgery_data =surgery_details_data(academic_year,q_year,partner_id,donor_id,district_id)
        pt_heading,a_data,t_data,state_dist_name = program_tracker_data(academic_year,q_year,partner_id,donor_id,district_id)
        return  export_excel_donor(child_scr_data,spec_comp_data,teach_train_data,surgery_data,academic_year,q_year,pt_heading,a_data,t_data,state_dist_name)
    return render(request, 'custom_report/performance_tracker.html', locals())


def child_screening_data(academic_year,q_year,partner_id,donor_id,district_id):
    where_cond = ''
    # if partner_id != '':
    #     where_cond += f" and partner_id = {int(partner_id)}"
    # if donor_id != '':
    #     where_cond += f" and donor_id = {int(donor_id)}"
    if district_id != '':
        where_cond += f" and school_district_id = {int(district_id)}"

    sql_query = f"""WITH numbered_rows AS (
                    SELECT 
                        row_number() OVER () AS row_num,
                        month,
                        student_name,
                        age,
                        gender,
                        school_type,
                        school_name,
                        school_district,
                        school_state,
                        remarks,
                        screening,
                        spec_prescriped,
                        spec_issued,
                        surgery_advised,
                        surgery_provided
                    FROM student_screening_data
                    WHERE p_quarter = '{q_year}' AND p_academic_year = '{academic_year}'   {where_cond}
                )
                SELECT jsonb_agg(
                    jsonb_build_array(
                        row_num::text,
                        month::text,
                        student_name::text,
                        age::text,
                        gender::text,
                        school_type::text,
                        school_name::text,
                        school_district::text,
                        school_state::text,
                        remarks::text,
                        screening::text,
                        spec_prescriped::text,
                        spec_issued::text,
                        surgery_advised::text,
                        surgery_provided::text
                    )
                )::text
                FROM numbered_rows;

            """
    data = return_sql_results_json(sql_query)
    return data

def surgery_details_data(academic_year,q_year,partner_id,donor_id,district_id):
    where_cond = ''
    # if partner_id != '':
    #     where_cond += f" and partner_id = {int(partner_id)}"
    # if donor_id != '':
    #     where_cond += f" and donor_id = {int(donor_id)}"
    if district_id != '':
        where_cond += f" and school_district_id = {int(district_id)}"
    sql_query = f"""
            WITH numbered_rows AS (
                    SELECT 
                        row_number() OVER () AS row_num,
                        student_name,
                        age,
                        gender,
                        school_name,
                        school_district,
                        school_state,
                        type_of_surgery,
                        date_of_surgery,
                        eye_operated_upon,
                        phase
                    FROM surgery_data
                    WHERE p_quarter = '{q_year}' AND p_academic_year = '{academic_year}' {where_cond}
                )
                SELECT jsonb_agg(
                    jsonb_build_array(
                        row_num::text,
                        student_name,
                        age,
                        gender,
                        school_name,
                        school_district,
                        school_state,
                        type_of_surgery,
                        date_of_surgery,
                        eye_operated_upon,
                        phase
                    )
                )::text
                FROM numbered_rows;

            """
    data = return_sql_results_json(sql_query)
    return data

def program_tracker_data(academic_year,q_year,partner_id,donor_id,district_id):
    heading = [f'Cumulative Target till date (June to March {academic_year.split("-")[1]})',f'Actual s till date (June to March {academic_year.split("-")[1]})','% Achieved',f'July {academic_year.split("-")[0]}- September {academic_year.split("-")[0]}',f'October {academic_year.split("-")[0]} - December {academic_year.split("-")[0]}',f'January {academic_year.split("-")[1]} - March {academic_year.split("-")[1]}',f'April {academic_year.split("-")[1]} - June {academic_year.split("-")[1]}']
    where_cond,state_dist_name = '',''
    if district_id != '':
        where_cond += f" and boundary_district_id = {int(district_id)}"
        state_dist_name_query = f"""
        select t2.name||' - '||t1.name from application_master_boundary t1 inner join application_master_boundary t2 on t1.parent_id = t2.id where t1.id = {int(district_id)}
                """
        state_dist_name = return_sql_results(state_dist_name_query)
    a_sql_query  = f"""
            WITH target_data AS (
            SELECT 
            DISTINCT ON (quarterly_year)
            quarterly_year,
            children_screened_q_target,
            children_prescribed_spectacles_q_target,
            children_provided_spectacles_q_target,
            children_refered_surgery_q_target,
            children_provided_surgery_q_target,
            schools_covered_q_target,
            teacher_trained_q_target,
            spectacle_compliance_3_months_q_target
        FROM 
            program_tracker_data
        where financial_year = '{academic_year}' {where_cond}
        ORDER BY 
            quarterly_year
            ),summed_data AS (
                    SELECT 
                        t1.quarterly_year,
                        t2.children_screened_q_target as children_screened_q_target,
                        t2.children_prescribed_spectacles_q_target as children_prescribed_spectacles_q_target,
                        t2.children_provided_spectacles_q_target as children_provided_spectacles_q_target,
                        t2.children_refered_surgery_q_target as children_refered_surgery_q_target,
                        t2.children_provided_surgery_q_target as children_provided_surgery_q_target,
                        t2.schools_covered_q_target as schools_covered_q_target,
                        t2.teacher_trained_q_target as teacher_trained_q_target,
                        t2.spectacle_compliance_3_months_q_target as spectacle_compliance_3_months_q_target,
                        t1.children_screened_ach AS children_screened_ach,
                        t1.children_prescribed_spectacles_ach AS children_prescribed_spectacles_ach,
                        t1.children_provided_spectacles_ach AS children_provided_spectacles_ach,
                        t1.children_refered_surgery_ach AS children_refered_surgery_ach,
                        t1.children_provided_surgery_ach AS children_provided_surgery_ach,
                        t1.schools_covered_ach AS schools_covered_ach,
                        t1.teacher_trained_ach AS teacher_trained_ach,
                        t1.spectacle_compliance_3_months_ach AS spectacle_compliance_3_months_ach,
                        sum(coalesce(t1.children_screened_male,0)) as children_screened_male,
                        sum(coalesce(t1.children_screened_female,0)) as children_screened_female,
                        sum(coalesce(t1.children_screened_total,0)) as children_screened_total,
                        SUM(COALESCE(t1.children_prescribed_spectacles_male, 0)) AS children_prescribed_spectacles_male,
                        SUM(COALESCE(t1.children_prescribed_spectacles_female, 0)) AS children_prescribed_spectacles_female,
                        SUM(COALESCE(t1.children_prescribed_spectacles_total, 0)) AS children_prescribed_spectacles_total,
                        SUM(COALESCE(t1.children_provided_spectacles_male, 0)) AS children_provided_spectacles_male,
                        SUM(COALESCE(t1.children_provided_spectacles_female, 0)) AS children_provided_spectacles_female,
                        SUM(COALESCE(t1.children_provided_spectacles_total, 0)) AS children_provided_spectacles_total,
                        SUM(COALESCE(t1.children_refered_surgery_male, 0)) AS children_refered_surgery_male,
                        SUM(COALESCE(t1.children_refered_surgery_female, 0)) AS children_refered_surgery_female,
                        SUM(COALESCE(t1.children_refered_surgery_total, 0)) AS children_refered_surgery_total,
                        SUM(COALESCE(t1.children_provided_surgery_male, 0)) AS children_provided_surgery_male,
                        SUM(COALESCE(t1.children_provided_surgery_female, 0)) AS children_provided_surgery_female,
                        SUM(COALESCE(t1.children_provided_surgery_total, 0)) AS children_provided_surgery_total,
                        SUM(COALESCE(t1.schools_covered_male, 0)) AS schools_covered_male,
                        SUM(COALESCE(t1.schools_covered_female, 0)) AS schools_covered_female,
                        SUM(COALESCE(t1.schools_covered_total, 0)) AS schools_covered_total,
                        SUM(COALESCE(t1.teacher_trained_male, 0)) AS teacher_trained_male,
                        SUM(COALESCE(t1.teacher_trained_female, 0)) AS teacher_trained_female,
                        SUM(COALESCE(t1.teacher_trained_total, 0)) AS teacher_trained_total,
                        SUM(COALESCE(t1.spectacle_compliance_3_months_male, 0)) AS spectacle_compliance_3_months_male,
                        SUM(COALESCE(t1.spectacle_compliance_3_months_female, 0)) AS spectacle_compliance_3_months_female,
                        SUM(COALESCE(t1.spectacle_compliance_3_months_total, 0)) AS spectacle_compliance_3_months_total 
                    FROM 
                        program_tracker_data t1 
                        left join target_data t2 on t1.quarterly_year = t2.quarterly_year
                        where financial_year = '{academic_year}' {where_cond}
                    GROUP BY 
                        t1.quarterly_year,
                        t2.children_screened_q_target,
                        t2.children_prescribed_spectacles_q_target,
                        t2.children_provided_spectacles_q_target,
                        t2.children_refered_surgery_q_target,
                        t2.children_provided_surgery_q_target,
                        t2.schools_covered_q_target,
                        t2.teacher_trained_q_target,
                        t2.spectacle_compliance_3_months_q_target,
                        t1.children_screened_ach,
                        t1.children_prescribed_spectacles_ach,
                        t1.children_provided_spectacles_ach,
                        t1.children_refered_surgery_ach,
                        t1.children_provided_surgery_ach,
                        t1.schools_covered_ach,
                        t1.teacher_trained_ach,
                        t1.spectacle_compliance_3_months_ach
                )
                SELECT 
                    quarterly_year,
                    json_build_array(
                        json_build_array(
                        children_screened_q_target,
                            children_screened_male,
                            children_screened_female,
                            children_screened_total,
                            children_screened_ach
                        ),
                        json_build_array(
                            children_prescribed_spectacles_q_target,
                            children_prescribed_spectacles_male,
                            children_prescribed_spectacles_female,
                            children_prescribed_spectacles_total,
                            children_prescribed_spectacles_ach
                        ),
                        json_build_array(
                            children_provided_spectacles_q_target,
                            children_provided_spectacles_male,
                            children_provided_spectacles_female,
                            children_provided_spectacles_total,
                            children_provided_spectacles_ach
                        ),
                        json_build_array(
                            children_refered_surgery_q_target,
                            children_refered_surgery_male,
                            children_refered_surgery_female,
                            children_refered_surgery_total,
                            children_refered_surgery_ach
                        ),
                        json_build_array(
                            children_provided_surgery_q_target,
                            children_provided_surgery_male,
                            children_provided_surgery_female,
                            children_provided_surgery_total,
                            children_provided_surgery_ach
                        ),
                        json_build_array(
                            schools_covered_q_target,
                            schools_covered_male,
                            schools_covered_female,
                            schools_covered_total,
                            schools_covered_ach
                        ),
                        json_build_array(
                        teacher_trained_q_target,
                            teacher_trained_male,
                            teacher_trained_female,
                            teacher_trained_total,
                            teacher_trained_ach
                        ),
                        json_build_array(
                        spectacle_compliance_3_months_q_target,
                            spectacle_compliance_3_months_male,
                            spectacle_compliance_3_months_female,
                            spectacle_compliance_3_months_total,
                            spectacle_compliance_3_months_ach 
                        )
                    ) as aggregated_data
                FROM 
                    summed_data
                ORDER BY 
                    CASE quarterly_year
                        WHEN 'Q1' THEN 1
                        WHEN 'Q2' THEN 2
                        WHEN 'Q3' THEN 3
                        WHEN 'Q4' THEN 4
                        ELSE 5
                    END;
            """

    t_sql_query = f"""
            with ach_data as (
            select financial_year,512 as i_id,sum(children_screened_total) as cum_total ,sum(children_screened_ach) as cum_ach
            from program_tracker_data t1 where t1.financial_year = '{academic_year}' {where_cond} group by financial_year
            union ALL
            select financial_year,513 as i_id,sum(children_prescribed_spectacles_total) as cum_total ,sum(children_prescribed_spectacles_ach) as cum_ach
            from program_tracker_data t1 where t1.financial_year = '{academic_year}' {where_cond} group by financial_year
            union ALL
            select financial_year,514 as i_id,sum(children_provided_spectacles_total) as cum_total ,sum(children_provided_spectacles_ach) as cum_ach
            from program_tracker_data t1 where t1.financial_year = '{academic_year}' {where_cond} group by financial_year
            union ALL
            select financial_year,515 as i_id,sum(children_refered_surgery_total) as cum_total ,sum(children_refered_surgery_ach) as cum_ach
            from program_tracker_data t1 where t1.financial_year = '{academic_year}' {where_cond} group by financial_year
            union ALL
            select financial_year,516 as i_id,sum(children_provided_surgery_total) as cum_total ,sum(children_provided_surgery_ach) as cum_ach
            from program_tracker_data t1 where t1.financial_year = '{academic_year}' {where_cond} group by financial_year
            union ALL
            select financial_year,517 as i_id,sum(schools_covered_total) as cum_total ,sum(schools_covered_ach) as cum_ach
            from program_tracker_data t1 where t1.financial_year = '{academic_year}' {where_cond} group by financial_year
            union ALL
            select financial_year,518 as i_id,sum(teacher_trained_total) as cum_total ,sum(teacher_trained_ach) as cum_ach
            from program_tracker_data t1 where t1.financial_year = '{academic_year}' {where_cond} group by financial_year
            union ALL
            select financial_year,519 as i_id,sum(spectacle_compliance_3_months_total) as cum_total ,sum(spectacle_compliance_3_months_ach) as cum_ach
            from program_tracker_data t1 where t1.financial_year = '{academic_year}' {where_cond} group by financial_year 
        )
        select t2.name,
        t1.annual_target,
        sum(t1.q1_target+t1.q2_target+t1.q3_target+t1.q4_target),
        coalesce(t3.cum_total,0),
        coalesce(t3.cum_ach,0)
        from reports_quarterlyreport t1 
        left join application_master_masterlookup t2
        on t1.indicator_id = t2.id
        left join ach_data t3 on t3.i_id = t1.indicator_id
        where t1.academic_year = '{int(academic_year.split("-")[0])}'
        group by t2.name,
        t1.annual_target,
        t3.cum_total,
        t3.cum_ach,
        t2.listing_order
        order by t2.listing_order
    
            """
    a_data = return_sql_results(a_sql_query)
    t_data = return_sql_results(t_sql_query)
    return heading,a_data,t_data,state_dist_name


def spec_compliance_data(academic_year,q_year,partner_id,donor_id,district_id):
    where_cond = ''
    # if partner_id != '':
    #     where_cond += f" and partner_id = {int(partner_id)}"
    # if donor_id != '':
    #     where_cond += f" and donor_id = {int(donor_id)}"
    if district_id != '':
        where_cond += f" and school_district_id = {int(district_id)}"
    sql_query = f"""
                    WITH numbered_rows AS (
                    SELECT 
                        row_number() OVER () AS row_num,
                        student_name,age,gender,school_name,school_state,spectacle_provided_on,using_spectacles,waering_complaince_after_3_month,reason_for_not_wearing
                    FROM spectacle_complaince_data 
                    WHERE 1=1 and p_quarter = '{q_year}' AND p_academic_year = '{academic_year}'
                    {where_cond}
                )
                SELECT jsonb_agg(
                    jsonb_build_array(
                        row_num::text,
                        student_name,age,gender,school_name,school_state,spectacle_provided_on,using_spectacles,waering_complaince_after_3_month,reason_for_not_wearing
                    )
                )::text
                FROM numbered_rows;
            """
    data = return_sql_results_json(sql_query)
    return data

def teacher_trained_data(academic_year,q_year,partner_id,donor_id,district_id):
    where_cond = ''
    # if partner_id != '':
    #     where_cond += f" and partner_id = {int(partner_id)}"
    # if donor_id != '':
    #     where_cond += f" and donor_id = {int(donor_id)}"
    if district_id != '':
        where_cond += f" and school_district_id = {int(district_id)}"
    sql_query = f"""
                    WITH numbered_rows AS (
                    SELECT 
                        row_number() OVER () AS row_num,
                        school_name,school_district,school_state,month_of_training,teacher_name,gender,trained
                    FROM teacher_training_data 
                    WHERE 1=1 and p_quarter = '{q_year}' AND p_academic_year = '{academic_year}' {where_cond}
                )
                SELECT jsonb_agg(
                    jsonb_build_array(
                        row_num::text,
                        school_name,school_district,school_state,month_of_training,teacher_name,gender,trained
                    )
                )::text
                FROM numbered_rows;
            """
    data = return_sql_results_json(sql_query)
    return data


def export_excel_donor(child_screening_data,spectacle_comp_data,teacher_training_data,surgery_details_data,financial_year,quarterly_year,pt_heading,a_data,t_data,state_dist_name):
    excel_path = settings.PERFORMANCE_TRACKER_TEMPLATES + '/performance-tracker-template.v.0.2.xlsx'  
    workbook = openpyxl.load_workbook(f'{excel_path}')
    worksheet_child_program_tracker = workbook.worksheets[0] 
    worksheet_child_screening = workbook.worksheets[1] 
    worksheet_surgery = workbook.worksheets[2] 
    worksheet_teacher_training = workbook.worksheets[3] 
    worksheet_spec_complaince = workbook.worksheets[4] 
    child_scr_data = child_screening_data if child_screening_data else [['No Data Avilable in the Financial Year and Quarterly Year']]
    teach_train_data = teacher_training_data if teacher_training_data else [['No Data Avilable in the Financial Year and Quarterly Year']]
    spec_comp_data = spectacle_comp_data if spectacle_comp_data else [['No Data Avilable in the Financial Year and Quarterly Year']]
    surgery_data = surgery_details_data if surgery_details_data else [['No Data Avilable in the Financial Year and Quarterly Year']]
    program_tracker_a_data = a_data if a_data else [['No Data Avilable in the Financial Year and Quarterly Year']]
    # sheet 1 - Program Tracker
    # import ipdb;ipdb.set_trace()
    #append heading
    start_row = 2
    for col_num, cell_value in enumerate(pt_heading[:3], start=1):
        worksheet_child_program_tracker.cell(row=start_row, column=3+col_num, value=cell_value)
    start_row = 3
    for col_num, cell_value in enumerate(pt_heading[3:], start=1):
        col_idx = 7+(5*(col_num-1))
        worksheet_child_program_tracker.cell(row=start_row, column=col_idx, value=cell_value)
    
    if state_dist_name:
        worksheet_child_program_tracker.cell(row=5, column=1, value=state_dist_name[0][0])
    # import ipdb;ipdb.set_trace()
    if t_data:
        start_row = 5
        for row_idx, row_data in enumerate(t_data):
            for idx,data in enumerate(row_data,start=1):
                col_idx = idx + 1
                worksheet_child_program_tracker.cell(row=start_row, column=col_idx, value=data)
            start_row += 1
    # Define a dictionary to map 'Q1', 'Q2', 'Q3', and 'Q4' to their respective column offsets
    quarter_to_column_offset = {
        'Q1': 6,
        'Q2': 11,
        'Q3': 16,
        'Q4': 21
    }
    # Iterate over program tracker data and fill in the cells
    for row_data in program_tracker_a_data:
        quarter = row_data[0]
        if quarter in quarter_to_column_offset:
            col_offset = quarter_to_column_offset[quarter]
            for idx, data_list in enumerate(row_data[1], start=1):
                start_row = 4 + idx
                for col_num, cell_value in enumerate(data_list, start=1):
                    worksheet_child_program_tracker.cell(row=start_row, column=col_offset + col_num, value=cell_value)
    #sheet2 - Children Screening Data
    worksheet_child_screening['A1'] = f"Children Screening Details - {financial_year} - {quarterly_year}"
    start_row = 3
    for row_data in child_screening_data:
        for col_num, cell_value in enumerate(row_data, start=1):
            worksheet_child_screening.cell(row=start_row, column=col_num, value=cell_value)
        start_row += 1
    #sheet3 - Surgery Details
    start_row = 3
    for row_data in surgery_data:
        for col_num, cell_value in enumerate(row_data, start=1):
            worksheet_surgery.cell(row=start_row, column=col_num, value=cell_value)
        start_row += 1
    #sheet 4 - Teacher Training
    start_row = 3
    for row_data in teach_train_data:
        for col_num, cell_value in enumerate(row_data, start=1):
            worksheet_teacher_training.cell(row=start_row, column=col_num, value=cell_value)
        start_row += 1
    #sheet5 - Spectacles Complaince
    start_row = 2
    for row_data in spec_comp_data:
        for col_num, cell_value in enumerate(row_data, start=1):
            worksheet_spec_complaince.cell(row=start_row, column=col_num, value=cell_value)
        start_row += 1
    folder_file_name = f"PERFORMANCE_TRACKER_REPORT-{financial_year}-{quarterly_year}_{datetime.today().strftime('%d%m%y%H%M')}.xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT + '/temp_donor_report/', folder_file_name)
    workbook.save(file_path)
    if os.path.exists(file_path):
        with open(file_path, "rb") as excel:
            data = excel.read()
            response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={folder_file_name}'
            os.remove(file_path)
    return response


def generate_export_excel(report_title, data_query_list, table_headers, custom_export_headers, sheet_names):
    from django.conf import settings
    excelWriter = None
    try:
        # Excel path
        rows_in_chunk = 20000
        db_name = settings.DATABASES['default'].get('NAME')
        username = settings.DATABASES['default'].get('USER')
        password = settings.DATABASES['default'].get('PASSWORD')
        hostname = settings.DATABASES['default'].get('HOST')
        # IMPORTANT : Please ensure @ not used in the username and password. This will affect the connection string
        conn_str = "postgresql+psycopg2://" + username + ":" + password + \
            "@" + hostname + "/" + db_name + "?client_encoding=UTF8"
        MEDIA_ROOT = str(settings.MEDIA_ROOT) + '/temp_export_data/'
        import datetime
        file_name = re.sub("(\s)|(')|(-)", '_', report_title)
        folder_file_name = file_name + "_" + \
            datetime.datetime.today().strftime("%d%m%y%H%M%S%f") + ".xlsx"
        attachment_name = file_name + "_" + \
            datetime.datetime.today().strftime("%d%m%y%H%M") + ".xlsx"
        excelWriter = pd.ExcelWriter(MEDIA_ROOT+folder_file_name, mode='w')
        for idx, sql_query in enumerate(data_query_list):
            # sheetname is limited to 30 chars othewise excel gives an error while opening
            write_to_excel_from_normalized_table(
                conn_str, sql_query, table_headers[idx], custom_export_headers[idx], rows_in_chunk, sheet_names[idx][0:30], excelWriter)

    finally:
        if excelWriter:
            excelWriter.close()

    if os.path.exists(MEDIA_ROOT+folder_file_name):
        excel = open(MEDIA_ROOT+folder_file_name, "rb")
        data = excel.read()
        response = HttpResponse(
            data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % attachment_name
        os.remove(MEDIA_ROOT+folder_file_name)
        return response


def get_loc_filter_value(user_filter_value, user_location_data_value):
    """
    Returns the location filter value based on the provided user filter value and user location data value. 
    """
    loc_id = ''
    if user_filter_value != '':
        loc_id = user_filter_value
    elif user_location_data_value != 0:
        loc_id = str(user_location_data_value)
    return loc_id


def update_location_data_with_user_filter(model, loc_data_index, selected_loc_id, user_location_data):
    """
    Updates the user location data with the selected location ID and model-specific data.
    """
    loc_data = user_location_data[loc_data_index]
    if loc_data[2] == False:
        if model == MissionIndicatorCategory:
            loc_list = list(MissionIndicatorCategory.objects.filter(
                active=2).values_list('id', 'name').order_by('name'))
            loc_data[4].update({str(0): loc_list})
        elif model == MissionIndicator and user_location_data[1][0] != '0':
            loc_list = list(MissionIndicator.objects.filter(
                category_id=user_location_data[0][0], active=2).values_list('id', 'name').order_by('name'))
            loc_data[4].update({str(user_location_data[0][0]): loc_list})
    user_location_data[loc_data_index] = loc_data
    user_location_data[loc_data_index][0] = int(
        selected_loc_id) if selected_loc_id != '' else 0
    return user_location_data


def get_filter_data(request, req_data, f_info,mission_id):
    import dateutil.relativedelta
    filter_values = []
    quarter_month_mapper = {"1": 4, "2": 4, "3": 4, "4": 1, "5": 1,
                            "6": 1, "7": 2, "8": 2, "9": 2, "10": 3, "11": 3, "12": 3}
    current_month = dt.now().strftime('%Y%-m')
    current_qtr = quarter_month_mapper.get(current_month[4:])
    current_year = int(current_month[0:4])
    current_year = current_year - 1 if current_qtr == 4 else current_year
    extended_filters_dict = {}
    # prepare the filter values with fixed order - state, district, shelter and rest of the filters added at the end
    filter_display_order = -1
    f_labels = f_info['filter_labels']
    filter_keys = f_info['filter_labels'].keys()
    user_filter_values = {}
    display_order = f_info['display_order']
    for key in filter_keys:
        str_val = req_data.get(key, '')
        if key == 'start_month':
            str_val=str_val.replace('-','')
        elif key == 'fv_start_month_year':
            str_val=str_val.replace('-','')
        elif key == 'fv_end_month_year':
            str_val=str_val.replace('-','')
        elif key == 'fv_month_year' :
            if str_val == '':
                previous_month=dt.now() + dateutil.relativedelta.relativedelta(months=-1)
                previous_year_month=previous_month.strftime('%Y%-m')
                current_filtered_qtr = quarter_month_mapper.get(previous_year_month[4:])
                selected_year = int(previous_year_month[0:4])
                selected_year = selected_year - 1 if current_filtered_qtr == 4 else selected_year
                str_val=previous_month.strftime('%Y%#m')
            else:
                str_val=str_val.replace('-','')
                current_filtered_qtr = quarter_month_mapper.get(str_val[4:].lstrip('0'))
                selected_year = int(str_val[0:4])
                selected_year = selected_year - 1 if current_filtered_qtr == 4 else selected_year
            user_filter_values.update({'fv_fy_start_month_year': str(selected_year)+'04'})
            user_filter_values.update({'fv_fy_year': str(selected_year)})
        elif request.method == 'POST' and key == 'partner' and str_val == '' :
            str_val=str(request.session['user_partner_list'])[1:-1]
        elif request.method == 'POST' and key == 'project' and str_val == '' :
            str_val=str(request.session['user_project_list'])[1:-1]
        user_filter_values.update({key: str_val})
    user_location_data = request.session['user_location_data'] if 'user_location_data' in request.session else None
    if user_location_data == None:
        load_user_details_to_sessions(request)
        user_location_data = request.session['user_location_data']
    user_location_data = copy.deepcopy(user_location_data)
    category_id = get_loc_filter_value(user_filter_values.get(
        'category', ''), user_location_data[0][0])
    indicator_id = get_loc_filter_value(user_filter_values.get(
        'indicator', ''), user_location_data[1][0])
    
    user_location_data = update_location_data_with_user_filter(
        MissionIndicatorCategory, 0, category_id, user_location_data)
    user_location_data = update_location_data_with_user_filter(
        MissionIndicator, 1, indicator_id, user_location_data)
    user_location_dict = request.session['user_location_dict'] if 'user_location_dict' in request.session else None
    loc_data = None
    for i in display_order:
        filter_values.append([])
    for k in filter_keys:
        data_list = []
        filter_display_order = -1
        if k in display_order:
            filter_display_order = display_order.index(k)

        if k == 'donor':
            donor_list = Donor.objects.filter(id__in=request.session['user_donor_list'],active=2).values_list('id', 'name')
            data_list = [(str(item[0]), item[1])
                         for item in donor_list]
            data_id = user_filter_values.get('donor', '')
            filter_type = 'select'
        elif k == 'partner':
            partner_list = Partner.objects.filter(id__in=request.session['user_partner_list'],active=2).values_list('id', 'name')
            data_list = [(str(item[0]), item[1])
                         for item in partner_list]
            data_id = user_filter_values.get('partner', '')
            filter_type = 'select'
        elif k == 'project':
            project_list = Project.objects.filter(id__in=request.session['user_project_list'],partner_mission_mapping__mission_id=mission_id,active=2).values_list('id', 'name')
            data_list = [(str(item[0]), item[1])
                         for item in project_list]
            data_id = user_filter_values.get('project', '')
            filter_type = 'select'
        elif k == 'category':
            loc_data = user_location_data[0][4]
            data_id = str(user_location_data[0][0])
            data_list = loc_data.get('0', [])
            if (data_id == '0' or len(data_list) == 1) and user_location_data[0][2] == True:
                query_data_id = user_location_dict.get('category', [])
                query_data_id = ','.join([str(i) for i in query_data_id])
                user_filter_values.update({'category': query_data_id})
            filter_type = 'select'
        elif k == 'indicator':
            loc_data = user_location_data[1][4]
            data_id = str(user_location_data[1][0])
            state_id = str(user_location_data[0][0])
            data_list = loc_data.get(state_id, [])
            if (data_id == '0' or len(data_list) == 1) and user_location_data[1][2] == True:
                query_data_id = user_location_dict.get('indicator', [])
                query_data_id = ','.join([str(i) for i in query_data_id])
                user_filter_values.update({'indicator': query_data_id})
            filter_type = 'select'
        elif k == 'start_month':
            data_list = []
            data_id = user_filter_values.get('start_month', '')
            if data_id != '':
                data_id='-'.join([data_id[:4],data_id[4:6]])
            filter_type = 'month'
        elif k == 'fv_month_year':
            data_list = []
            data_id = user_filter_values.get('fv_month_year') if user_filter_values.get('fv_month_year') != '' else current_month
            if data_id != '' and '-' not in data_id:
                data_id='-'.join([data_id[:4],data_id[4:6]])
            filter_type = 'month'
        elif k == 'fv_start_month_year':
            data_list = []
            data_id = user_filter_values.get('fv_start_month_year') 
            if data_id != '':
                data_id='-'.join([data_id[:4],data_id[4:6]])
            filter_type = 'month'
        elif k == 'fv_end_month_year':
            data_list = []
            data_id = user_filter_values.get('fv_end_month_year') 
            if data_id != '':
                data_id='-'.join([data_id[:4],data_id[4:6]])
            filter_type = 'month'
        elif k == 'grading':
            data_list = [(" A ", " A "), (" B ", " B "), (" C "," C "), (" D ", " D "), (" E ", " E "),(" O ", " O ")]
            data_id = user_filter_values.get('grading', '')
            
            filter_type = 'select'
        elif k == 'case_worker_qtr_start':
            data_list = [("202102", "FY'21 Q2"), ("202103", "FY'21 Q3"), ("202104",
                                                                          "FY'21 Q4"), ("202201", "FY'22 Q1"), ("202202", "FY'22 Q2")]
            data_id = user_filter_values.get('case_worker_qtr_start', '')
            data_id = str(current_year*100 +
                          current_qtr) if data_id == '' else data_id
            filter_type = 'select'
            qtr_end = user_filter_values.get('case_worker_qtr_end', '')
            qtr_end = str(current_year*100 +
                          current_qtr) if qtr_end == '' else qtr_end
            extended_filters_dict = add_extended_filters(data_id, qtr_end)
        elif k == 'case_worker_qtr_end':
            data_list = [("202102", "FY'21 Q2"), ("202103", "FY'21 Q3"), ("202104",
                                                                          "FY'21 Q4"), ("202201", "FY'22 Q1"), ("202202", "FY'22 Q2")]
            data_id = user_filter_values.get('case_worker_qtr_end', '')
            data_id = str(current_year*100 +
                          current_qtr) if data_id == '' else data_id
            filter_type = 'select'
        if filter_display_order == -1:
            filter_values.append(
                [k, data_list, data_id, f_labels.get(k, ''), filter_type])
        else:
            filter_values[filter_display_order] = [
                k, data_list, data_id, f_labels.get(k, ''), filter_type]

    return user_location_data, filter_values, user_filter_values, extended_filters_dict


def add_extended_filters(qtr_start, qtr_end):
    import datetime
    qtr_start_months = {"1": 4, "2": 7, "3": 10, "4": 1}
    qtr_end_months = {"1": 6, "2": 9, "3": 12, "4": 3}
    extended_filters_dict = {}
    fy_start_year = qtr_start[:4]
    start_qtr = int(qtr_start[-2:])
    end_qtr = int(qtr_end[-2:])
    qtr_start_month = qtr_start_months.get(str(start_qtr))
    fy_start_year = int(fy_start_year) + \
        1 if qtr_start_month == 1 else int(fy_start_year)
    fy_end_year = int(qtr_end[:4])
    qtr_end_month = qtr_end_months.get(str(end_qtr))
    # generate the list of quearters to be included as per the start and end quarters
    qtr_range = []
    sq = start_qtr
    cy = fy_start_year
    while True:
        for i in range(sq, 5):
            qtr_range.append(str(cy*100+i))
            if cy == fy_end_year and i >= end_qtr:
                break
        if cy < fy_end_year:
            cy = cy + 1
            sq = 1
        else:
            break
    qtr_range_str = ''
    for idx, i in enumerate(qtr_range):
        if idx >= 1:
            qtr_range_str = qtr_range_str + ","
        qtr_range_str = qtr_range_str + "'" + i + "'"
    
    # caluclate the end and start months to be used in the query
    fy_end_year = int(
        fy_end_year) + 1 if qtr_end_month == 3 else int(fy_end_year)
    filter_start_date = str(fy_start_year)+"-" + \
        str(qtr_start_month).rjust(2, '0')+"-01"
    filter_end_date = str(fy_end_year)+"-" + \
        str(qtr_end_month).rjust(2, '0')+"-01"
    filter_end_date = datetime.datetime.strptime(filter_end_date, '%Y-%m-%d')
    filter_end_date = filter_end_date + relativedelta(months=1)
    filter_end_date = filter_end_date.strftime('%Y-%m-%d')
    qtr_range = []

    extended_filters_dict.update({'case_worker_qtr_ext_filter': """ and (case_worker_start_date < '""" + filter_end_date + """'::date 
                                and (case_worker_pause_date is null or case_worker_pause_date >= '""" + filter_start_date + """'::date)
                                and (case_worker_end_date is null 
                                    or case_worker_end_date >= '""" + filter_start_date + """'::date)
                                )"""})
    extended_filters_dict.update({'qtr_range_ext_filter': qtr_range_str})
    return extended_filters_dict


def update_variable_sort_details(sort_info, default_sort, location_name, location_filter, const_variable_location_name):
    # update sort info
    updated_sort_info = {}
    # update default sort
    if default_sort:
        default_sort_field = default_sort.get('sort_field', '')
        default_sort.update({'sort_field': default_sort_field.replace(
            const_variable_location_name, location_name)})
    return default_sort, updated_sort_info


def apply_variable_location_info(header, custom_export_header, data_query, count_query, sort_info, default_sort, user_filter_values, user_location_data):
    location_name = ''
    const_variable_location_name = '@@variable_location_name'
    location_filter = ''

    # update custom export header
    if custom_export_header:
        for row_idx, row in enumerate(custom_export_header):
            for col_idx, item in enumerate(row):
                if item == const_variable_location_name:
                    custom_export_header[row_idx][col_idx] = location_filter

    # update header data
    for row in header:
        for col in row:
            col_label = col.get('label')
            if col_label == const_variable_location_name:
                col_label = col_label.replace(
                    const_variable_location_name, location_filter)
                col.update({'label': col_label})
    default_sort, sort_info = update_variable_sort_details(
        sort_info, default_sort, location_name, location_filter, const_variable_location_name)
    # update data and count queries
    data_query = data_query.replace(
        const_variable_location_name, location_name)
    count_query = count_query.replace(
        const_variable_location_name, location_name)
    return header, custom_export_header, data_query, count_query, sort_info, default_sort


def set_sort_options(req_data, idx, s_info, default_sort):
    sort_field = req_data.get('sort_field_'+str(idx), '')
    sort_order = req_data.get('sort_order_'+str(idx), 'asc')
    # set the default sort order if configured in the report_query section - key : default_sort and value is a dict with sort_order and sort_fields
    # "default_sort":{"sort_field":"classification","sort_order":"asc"}
    if default_sort:
        if 'sort_field' in default_sort and sort_field == '':
            sort_field = default_sort.get('sort_field', '')
        if 'sort_order ' in default_sort and sort_field == '':
            sort_order = default_sort.get('sort_order', 'asc')
    elif sort_field == '':
        # when default sort not specified, set the sort field to first key in the list and order to asc
        for key in s_info:
            sort_field = s_info.get(key, '')
            sort_order = 'asc'
            break
    return sort_field, sort_order


def apply_filters_to_query(data_query, count_query, filter_info, sort_field, sort_order, user_filter_values, current_page, rows_per_page, extended_filter_dict, export_flag):
    # apply filters and sort conditions to query
    filter_cond = filter_info['filter_cond']
    # IMPORTANT - Please cast all timezone aware timestamp fields to timestamps without timezone in the query
    # or use the to_char function and pass required format
    for key in filter_cond.keys():
        filter_value = user_filter_values.get(key)
        updated_cond = ''
        if filter_value != None and filter_value != '' and filter_value != '0':
            f_cond = filter_cond[key]
            if f_cond.lower().replace(' ','').find("in(@@filter_value)") == -1:
                updated_cond = filter_cond[key].replace('@@filter_value',filter_value)
            else:
                updated_cond = filter_cond[key].replace(
                    '@@filter_value', filter_value)
        data_query = data_query.replace('@@'+key+'_filter', updated_cond)
        count_query = count_query.replace('@@'+key+'_filter', updated_cond)
    # apply extended filters dict
    for f_key, f_value in extended_filter_dict.items():
        data_query = data_query.replace('@@'+f_key, f_value)
        count_query = count_query.replace('@@'+f_key, f_value)
    limits_query = ''
    if export_flag is None or export_flag == False:
        limits_query = ' LIMIT ' + \
            str(rows_per_page) + ' OFFSET ' + \
            str(rows_per_page*(current_page-1))
    else:
        data_query = data_query.replace('<br/>', '')
    data_query = data_query.replace('@@LIMITS', limits_query)

    sortings = ''
    if sort_field != None and sort_field != '':
        sort_order = '' if sort_order == None else sort_order
        sortings = ' order by ' + sort_field + ' ' + sort_order + ' '
    data_query = data_query.replace('@@sortings', sortings)
    return data_query, count_query


def calculate_pagination_info(total_records, current_page, rows_per_page):
    display_page_range = 10
    num_pages = int(total_records/rows_per_page)
    num_pages = num_pages if total_records % rows_per_page == 0 else num_pages + 1
    start_page = (current_page - int(display_page_range/2))
    start_page = 1 if start_page < 1 else start_page
    end_page = start_page + display_page_range - 1
    end_page = num_pages if end_page > num_pages else end_page
    first_row = ((current_page-1)*rows_per_page) + 1
    last_row = (current_page*rows_per_page)
    last_row = total_records if last_row > total_records else last_row
    dicta = {"rec_count": total_records, "current_page": current_page, "rows_per_page": rows_per_page, "display_page_range": display_page_range,
             "num_pages": num_pages, "start_page": start_page, "end_page": end_page, "first_row": first_row, "last_row": last_row}
    return {"rec_count": total_records, "current_page": current_page, "rows_per_page": rows_per_page, "display_page_range": display_page_range,
            "num_pages": num_pages, "start_page": start_page, "end_page": end_page, "first_row": first_row, "last_row": last_row}


@ login_required(login_url='/login/')
def custom_report_reload(request, page_slug, report_slug):
    import sys
    import traceback
    html = ''
    rows_per_page = 10
    try:
        if request.method == 'POST':
            if request.method == "POST":
                req_data = request.POST
            elif request.method == "GET":
                req_data = request.GET
            # order of reports is specified in the ReportMeta default ordering
            # page_reports = get_list_or_404(ReportMeta, page_slug=page_slug, active=2)
            page_reports = ReportMeta.objects.filter(
                page_slug=page_slug, active=2).order_by('display_order')
            table_header = []
            report_slug_list = []
            data = []
            nowrap_cols = []
            user_sort_field = []
            user_sort_order = []
            sorting_field = []
            page_info = []
            user_filter_values = {}
            total_header_cols = []
            report_idx = 0
            for idx, report in enumerate(page_reports):
                if report.report_slug == report_slug:
                    report_idx = idx
                else:
                    continue
                report_slug_list.append(report.report_slug)
                r_query = report.report_query
                d_query = r_query['sql_query']
                c_query = r_query['count_query'] if r_query['count_query'] else ''
                mission_id = r_query['mission_id']
                f_info = report.filter_info
                s_info = report.sort_info
                e_header = report.custom_export_header
                headers = report.report_header  # table_header
                user_location_data, filter_values, user_filter_values, extended_filter_dict = get_filter_data(
                    request, req_data, f_info,mission_id)
                # update any variable_location_names  - in query, count query, sort and headers
                default_sort = r_query['default_sort'] if 'default_sort' in r_query else None
                headers, e_header, data_query, count_query, s_info, default_sort = apply_variable_location_info(
                    headers, e_header, d_query, c_query, s_info, default_sort, user_filter_values, user_location_data)
                sort_field, sort_order = set_sort_options(
                    req_data, report_idx, s_info, default_sort)
                user_sort_field.append(sort_field)
                user_sort_order.append(sort_order)

                current_page = int(req_data.get('page_'+str(report_idx), '0'))
                data_query, count_query = apply_filters_to_query(
                    data_query, count_query, f_info, sort_field, sort_order, user_filter_values, current_page, rows_per_page, extended_filter_dict, False)
                data.append(return_sql_results(data_query))

                # table header details
                table_header.append(headers)
                # get total columns count (colspan sum) - used to display the no records found row
                total_records = int(req_data.get(
                    'rec_count_'+str(report_idx), '0'))
                total_header_cols.append(total_records)

                p_info = calculate_pagination_info(
                    total_records, current_page, rows_per_page)
                pagination_range = range(p_info.get(
                    'start_page'), p_info.get('end_page')+1)
                page_info.append(p_info)
                sorting_field.append(s_info)
            html = render_to_string('reports/report_ajax_reload.html', {"req_data": req_data, "report_idx": report_idx, "table_header": table_header, "data": data,
                                                                        "nowrap_cols": nowrap_cols, "user_sort_field": user_sort_field, "user_sort_order": user_sort_order, "page_info": page_info,
                                                                        "pagination_range": pagination_range})
    except Exception as ex1:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        error_stack = repr(traceback.format_exception(
            exc_type, exc_value, exc_traceback))
    return HttpResponse(html)


def load_user_details_to_sessions(request):
    # Getting the user role config if not it will raise exception
    # [loc_id, "loc_name", loc_config, {loc_data}]
    # - config True if the location is from user location config, false if its a user selected location from filter dropdown
    # [district_id, "district_name", True, [district_id1, distrit_id2, district_id3], {"state_id":[(district_id1, "district_name1"),(district_id2, "district_name2")]}]
    try:

        # user_role_location_level_config = User.objects.get(
        #     id=request.user.id)
        # user_group = user_role_location_level_config.groups.first()
        # # content type id of the user location level - state/district/shelter
        # location_hierarchy_type_id = user_role_location_level_config.location_hierarchy_type.id
        user_based_categories =request.session['user_category_list']
        loc_ids = user_based_categories
        loc_id = ''
        loc_name = ''
        config_loc = False
        shelter_list, district_list, state_list = [], [], []
        shelter_data, district_data, state_data = {}, {}, {}
        shelter_config, district_config, state_config = False, False, False
        loc_list = MissionIndicatorCategory.objects.filter(id__in=user_based_categories).values_list(
            'id', 'name',).order_by('name')        
        state_config = True
        attrib_len = 2
        for item in loc_list:
            new_state = (item[0], item[1])
            state_list.append(new_state[0])
            s_list = state_data.get('0', [])
            s_list.append(new_state)
            state_data.update({'0': s_list})
            if attrib_len > 2:
                new_district = (item[2], item[3])
                district_list.append(new_district[0])
                district_for_state = district_data.get(str(new_state[0]), [])
                district_for_state.append(new_district)
                district_data.update({str(new_state[0]): district_for_state})
            if attrib_len > 4:
                new_shelter = (item[4], item[5])
                shelter_list.append(new_shelter[0])
                shelter_for_district = shelter_data.get(
                    str(new_district[0]), [])
                shelter_for_district.append(new_shelter)
                shelter_data.update(
                    {str(new_district[0]): shelter_for_district})
        user_location_data = []
        state_list = list(set(state_list))
        district_list = list(set(district_list))
        shelter_list = list(set(shelter_list))
        state_id, state_name = 0, ''
        if len(state_list) == 1:
            state_id = state_list[0]
            state_name = state_list[0]
        for loc_item in state_data:
            state_data.update(
                {loc_item: sorted(list(set(state_data.get(loc_item, []))), key=lambda x: x[1])})
        user_location_data.append(
            [state_id, state_name, state_config, state_list, state_data])
        district_id, district_name = 0, ''
        if len(district_list) == 1:
            district_id = district_list[0]
            district_name = district_list[0]
        for loc_item in district_data:
            district_data.update({loc_item: sorted(
                list(set(district_data.get(loc_item, []))), key=lambda x: x[1])})
        user_location_data.append(
            [district_id, district_name, district_config, district_list, district_data])
        shelter_id, shelter_name = 0, ''
        if len(shelter_list) == 1:
            shelter_id = shelter_list[0]
            shelter_name = shelter_list[0]
        for loc_item in shelter_data:
            shelter_data.update({loc_item: sorted(
                list(set(shelter_data.get(loc_item, []))), key=lambda x: x[1])})
        user_location_data.append(
            [shelter_id, shelter_name, shelter_config, shelter_list, shelter_data])
        user_location_dict = {'state': state_list,
                              'district': district_list, 'shelter': shelter_list}
    except User.DoesNotExist:
        configure_error = 'Username not configured . Please contact administration.'
        return configure_error

    request.session['user_location_data'] = user_location_data
    request.session['user_location_dict'] = user_location_dict


@ login_required(login_url='/login/')
def get_indicator(request):
    if request.method == 'GET' and request.is_ajax():
        selected_category = request.GET.get('selected_category', '')
        user_location_data = request.session['user_location_data'] if 'user_location_data' in request.session else None
        if user_location_data == None:
            load_user_details_to_sessions(request)
            user_location_data = request.session['user_location_data']
        loc_data = user_location_data[1]  # index 1 is category data
        result_set = []
        if loc_data[2] == True:  # user configured districts
            districts = loc_data[4].get(selected_category, [])
        else:  # all indicators for category
            districts = MissionIndicator.objects.filter(category_id=int(
                selected_category), active=2).order_by('name').values_list('id', 'name')
        for district in districts:
            result_set.append(
                {'id': district[0], 'name': district[1], })
        return HttpResponse(json.dumps(result_set))

@csrf_exempt
@ login_required(login_url='/login/')
def get_district(request):
    if request.method == 'POST':
        selected_projects = request.POST.get('selected_projects[]', '')
        district_ids = Project.objects.filter(id = int(selected_projects),active=2).values_list('district_id',flat=True) 
        districts = District.objects.filter(id__in=district_ids,active=2).order_by('name').values_list('id', 'name')
        result_set = []
        for district in districts:
            result_set.append(
                {'id': district[0], 'name': district[1], })
        return HttpResponse(json.dumps(result_set))


import json
from django.http import JsonResponse, HttpResponse

def get_project(request):
    if request.method == 'POST' and request.is_ajax():
        selected_partner = request.GET.get('selected_partner', '')
        if not selected_partner:
            return JsonResponse([], safe=False)
        try:
            selected_partner_id = int(selected_partner)
        except ValueError:
            return JsonResponse([], safe=False)
        user_location_data = request.session.get('user_location_data')
        if not user_location_data:
            load_user_details_to_sessions(request)
            user_location_data = request.session['user_location_data']
        loc_data = user_location_data[1] 
        result_set = []
        if loc_data[2]:  
            districts = loc_data[4].get(selected_partner_id, [])
        else:  
            pmm_ids = PartnerMissionMapping.objects.filter(partner_id=selected_partner_id).values_list('id', flat=True)
            projects = Project.objects.filter(partner_mission_mapping_id__in=pmm_ids, active=2)
            for project in projects:
                result_set.append({'id': project.id, 'name': project.name})
        return JsonResponse(result_set, safe=False)
    return HttpResponse(status=400)

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def get_school(request):
    if request.method == 'POST':
        selected_partner = request.POST.get('partner_id[]', '')
        if selected_partner == '':
            return JsonResponse([], safe=False)
        else:  
            result_set = []
            pmm_ids = PartnerMissionMapping.objects.filter(partner_id = selected_partner).values_list('id', flat=True)
            district_ids = Project.objects.filter(partner_mission_mapping_id__in=pmm_ids, active=2).values_list('district_id',flat=True)
            boundary_ids = Boundary.objects.filter(code__in=[str(dist) for dist in district_ids],boundary_level_type_id =2).values_list('id',flat=True)
            district_id_list = [i for i in boundary_ids]
            school_list = return_sql_results(f"""select creation_key,profile_view->'data'->>'School name' from survey_beneficiaryresponse where survey_id = 1 and address_2 in ({str(district_id_list)[1:-1]})""")
            for school in school_list:
                result_set.append({'id': school[0], 'name': school[1]})
        return JsonResponse(result_set, safe=False)
    return HttpResponse(status=400)

#################################### ADD Export Report ##################################################

def execute_query(conn, sql_query, query_type):
    
    #typ1 = insert
    #typ2 = update
    #typ3 = select
    #typ4 = stored procedure
    #typ5 = insert returning id
    #typ6 = DDL
    # import ipdb;ipdb.set_trace()
    result = None
    cursor = None
    try:
        db_conn = None
        if conn is None:
            db_name = settings.DATABASES['default'].get('NAME')
            username = settings.DATABASES['default'].get('USER')
            password = settings.DATABASES['default'].get('PASSWORD')
            host = settings.DATABASES['default'].get('HOST') if settings.DATABASES['default'].get('HOST') else 'localhost'
            db_conn = psycopg2.connect(database=db_name, user=username, password=password, host=host)
        else:
            db_conn = conn
        cursor = db_conn.cursor()
        cursor.execute(sql_query)
        if query_type == 3 or query_type == 4:
            result = cursor.fetchall()
        elif query_type == 5:
            result = cursor.fetchone()[0]
    finally :
        if cursor:
            cursor.close()
        #close db_conn (psycopg2 conn) created if connection not passed as parametner
        if conn is None and db_conn:
            db_conn.close()
    return result

def get_session_name(request):
    menuobj = None
    try:
        if request.path == "/budget/manage/lineitem/" and request.GET.get('key') == "achievements":
            request.session['side_menu'] = "achievements-lineitem"
            request.session['parent_menu'] = "fund-management"
        elif request.path == "/budget/manage/lineitem/" and request.GET.get('key') == "planning":
            request.session['side_menu'] = "target-entry"
            request.session['parent_menu'] = "fund-management"
        else:
            try:
                menuobj = Menus.objects.get(backend_link__icontains = request.get_full_path())
            except:
                menuobj = Menus.objects.get_or_none(backend_link__icontains = request.path)
            # for child of chidl menu
            if menuobj.parent.parent:
                request.session['parent_menu'] = menuobj.parent.parent.slug if menuobj.parent.parent else menuobj.parent.slug
                request.session['parent_side_menu'] = menuobj.parent.slug if menuobj.parent else menuobj.slug
#...........for child menu
            elif menuobj.parent:
                if request.session.get('parent_side_menu'):
                    request.session.pop('parent_side_menu')
                request.session['parent_menu'] = menuobj.parent.slug if menuobj.parent else menuobj.slug
            request.session['side_menu'] = menuobj.slug if menuobj else ''
                
    except:
        pass
    return menuobj

month_dict={'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun','07':'Jul','08':'Aug', '09':'Sep','10':'Oct','11':'Nov','12':'Dec'}

def user_projects(user):
    user_projects = Project.objects.filter(active=2)
    userrole = UserRoles.objects.get_or_none(user=user)
    project_level = userrole.organization_unit.is_project_level if userrole and userrole.organization_unit else None
    if not user.is_superuser and project_level == True:
        user_role = get_role(user)
        user_project_relation = ProjectUserRelation.objects.filter(
            user__in=[user_role]).values_list('project', flat=True)
        user_projects = Project.objects.filter(
            id__in=user_project_relation, active=2)
    return user_projects

def get_report_query(request, widget_obj, page_id,rows_per_page, locations_list, filter_from_date, filter_to_date, location_id=None):
    import datetime
    import dateutil.relativedelta
    obj = eval(widget_obj.widgetquery)
    sql_query = obj.get('sql_query')
    # config_x_days = user_setup().get('config_x_days')
    filter_from_date_where,query, filter_to_date_where = '', '',''
    # date format - 'yyyy-mm-dd'
    # if filter_to_date and filter_from_date:
    #     filter_from_date_where = " and filter_date >= '" + filter_from_date + "' "
    #     filter_to_date_where = " and filter_date <= '" + filter_to_date + "' "
    # sql_query = sql_query.replace("@@filter_from_date", filter_from_date_where);
    # sql_query = sql_query.replace("@@filter_to_date", filter_to_date_where);
    district_ids = locations_list
    # village_id = locations_list[4] if len(locations_list)>=5 else 0
    # cluster_id = locations_list[3]  if len(locations_list)>=4 else 0
    # taluka_id = locations_list[2] if len(locations_list)>=3 else 0
    # district_id = locations_list[1] if len(locations_list)>=2 else 0
    # state_id = locations_list[0] if len(locations_list)>=1 else 0
    # village_filter,cluster_filter,taluka_filter,district_filter,state_filter = "","","","",""
    district_filter = ''
    if district_ids:
        district_filter = f" and address_2 in ({str(district_ids)[1:-1]})"
    sql_query = sql_query.replace("@@location_filter", district_filter)
    return sql_query

def load_user_details(request):
    user = request.user
    partner_ids = request.GET.getlist('partner','')
    school_ids = request.GET.getlist('school','')
    if user.groups.filter(name__in = ['Partner Data Entry Operator','Partner Admin','Project In-charge']).exists():# project incharge
        user_project=UserProjectMapping.objects.filter(user=request.user,active=2)
        user_project_ids = user_project.values_list('project__id',flat=True)
        user_partner_id = user_project.values_list('project__partner_mission_mapping__partner_id',flat=True)
        user_mission_id = user_project.values_list('project__partner_mission_mapping__mission_id',flat=True)
        user_donor_id=ProjectDonorMapping.objects.filter(project__id__in=user_project_ids,active=2).values_list('donor__id',flat=True).distinct()
        user_category_list=MissionIndicatorCategory.objects.filter(mission__id__in=user_mission_id,active=2).values_list('id',flat=True)
    elif user.is_superuser:
        user_mission_id=Mission.objects.filter(active=2).values_list('id',flat=True)
        user_project_ids=Project.objects.filter(active=2).values_list('id',flat=True)
        if partner_ids != '':
            user_project_ids=Project.objects.filter(active=2,partner_mission_mapping__partner_id__in = partner_ids).values_list('id',flat=True)
        user_partner_id=Partner.objects.filter(active=2).values_list('id',flat=True)
        user_donor_id=Donor.objects.filter(active=2).values_list('id',flat=True).distinct()
        user_category_list=MissionIndicatorCategory.objects.filter(active=2).values_list('id',flat=True)
    user_district_ids = Project.objects.filter(id__in = user_project_ids).values_list('district_id',flat=True)
    district_id_list = Boundary.objects.filter(code__in = [str(dist) for dist in user_district_ids],active=2,boundary_level_type_id=2).values_list('id',flat=True)
    request.session['user_mission_list']=list(user_mission_id)
    request.session['user_project_list']=list(user_project_ids)
    request.session['user_partner_list']=list(user_partner_id)
    request.session['user_donor_list']=list(user_donor_id)
    request.session['user_category_list']=list(user_category_list)
    request.session['user_district_list']=list(district_id_list)
    if school_ids != '':
        district_id_list = return_sql_results(f"""select address_2 from survey_beneficiaryresponse where creation_key in ({str(school_ids)[1:-1]}) and survey_id = 1 """)
        request.session['user_district_list']=[dist[0] for dist in district_id_list]
    return request






@login_required(login_url='/login/')
def export_reportcsv(request,slug):
    import re
    from django.db import connection
    from datetime import datetime
    menu_slug_name = get_session_name(request)
    today = datetime.today()
    today_month = month_dict['{:02d}'.format(today.month)]
    today_year = today.year
    filters_cond = False
    user_details = load_user_details(request)
    if 'user_partner_list' in request.session:
        partner_ids = request.session['user_partner_list']
    partnerlist = Partner.objects.filter(id__in=partner_ids).values("id","name")
    if 'user_district_list' in request.session:
        district_ids = request.session['user_district_list']
    # levels_to_filter = BoundaryLevel.objects.filter(code__gte=2, code__lte=2).order_by('code')
    location_id, from_date, to_date, frm, to = None, None, None, None, None
    filter_from_date = request.GET.get('from_date','') 
    filter_to_date = request.GET.get('to_date','')
    #frontend purpose we are adding one more varaibale
    search_filter_value = request.GET.get('search_box','')
    filter_locations = district_ids
    export_request = False
    media_url = settings.EXPORT_MEDIA_URL
    page_id, rows_per_page = 0, 10
    page_id = int(request.GET.get('page',0))
    if page_id <= 0:
        page_id = 1
    report_slug = slug
    user = request.user
    locations_list=[]
    if report_slug:
        widget_obj = DashboardChartWidgets.objects.get(slug = report_slug)
        qlist = eval(widget_obj.widgetquery).get('headers')
        if widget_obj.query_type == 'SQL_P':
            # sql_query = get_report_query_sp(request , widget_obj ,page_id,rows_per_page, filter_locations, filter_from_date, filter_to_date,location_id)
            sql_query = get_report_query(request , widget_obj ,page_id,rows_per_page, filter_locations, filter_from_date, filter_to_date,location_id)
            total_no_of_records = -1
            object_list = execute_query(connection, sql_query,typ=1,extra=True)
        else:
            if export_request == True:
                page_id = None
            sql_query = get_report_query(request , widget_obj ,page_id,rows_per_page, filter_locations, filter_from_date, filter_to_date,location_id)
            # get the count query and execute it to get the total records
            # This is executed every time so we dont pass the count in the query string and also
            # any additional records getting added can be included in the count
            if request.GET.get('partner'):
                partner_list_str = request.GET.getlist('partner')
                school_list_str = request.GET.getlist('school')
                partner_list_id = (str([eval(i) for i in partner_list_str]))[1:-1]
                filter_from_date = filter_to_date = ''
                # sql_query = get_report_query(request , widget_obj ,page_id,rows_per_page, filter_locations, filter_from_date, filter_to_date,location_id)
                # sql_query = sql_query.replace('@@searchfilter',f"and pp.id in ({partner_list_id}) and 1=1")
                url_filter_params = ""
                pmm_ids = PartnerMissionMapping.objects.filter(partner_id__in = [eval(i) for i in partner_list_str]).values_list('id', flat=True)
                district_ids = Project.objects.filter(partner_mission_mapping_id__in=pmm_ids, active=2).values_list('district_id',flat=True)
                boundary_ids = Boundary.objects.filter(code__in=[str(dist) for dist in district_ids],boundary_level_type_id =2).values_list('id',flat=True)
                district_id_list = [i for i in boundary_ids]
                school_list = return_sql_results(f"""select creation_key,profile_view->'data'->>'School name' from survey_beneficiaryresponse where survey_id = 1 and address_2 in ({str(district_id_list)[1:-1]}) """)
            search_filter_value = request.GET.get('search_box','')
            if search_filter_value != '':
                sql_query = sql_query.replace('@@searchfilter',"and upper(ss.name) like \'%"+search_filter_value.upper()+"%\' and 1=1")
            else:
                sql_query = sql_query.replace('@@searchfilter','and 1=1 ')
            count_sql_query = f'select count(*) from ({sql_query}) as x1'
            count_sql_query = re.sub(r'(LIMIT .*?$)',r') as x1',count_sql_query,1)
            total_no_of_records = execute_query(connection, count_sql_query,3)[0][0]
            total_pages =  (total_no_of_records//rows_per_page)+ 1 if total_no_of_records%rows_per_page != 0 else total_no_of_records/rows_per_page
            if (request.GET.get('Project','') != '' and widget_obj.id == 32) or widget_obj.id != 32: 
                object_list = execute_query(connection, sql_query,3)
            else:
                object_list = execute_query(connection, sql_query,3)
    object_list = list(map(lambda x : ['-' if i == None or i == '' else i for i in x], object_list))
    # skip pagination for stored procedure -- until the out param issue is fixed
    if total_no_of_records == -1:
        start_index = 1
        previous_page = 0
        next_page = 0
        total_pages = 1
        page_id = 1
    else:
        start_index = (page_id-1)*rows_per_page+1
        previous_page = page_id - 1
        next_page = page_id + 1
        total_pages =  (total_no_of_records//rows_per_page)+ 1 if total_no_of_records%rows_per_page != 0 else round(total_no_of_records/rows_per_page)
    return render(request,'reports/export_reportcsv.html',locals())



import xlsxwriter

def custom_report_csv(request, report_id):
    report_meta = ReportMeta.objects.get(id=report_id, active=2)
    sql_query = report_meta.report_query
    data = return_sql_results(sql_query['sql_query'])
    workbook = xlsxwriter.Workbook(f"{BASE_DIR+'/media/'+report_meta.report_slug}.xlsx")
    worksheet = workbook.add_worksheet()
    headers = report_meta.custom_export_header
    for col_num, header in enumerate(headers["headers"]):
        worksheet.write(0, col_num, header)
    # Write data to the Excel file
    for row_num, row_data in enumerate(data):
        for col_num, cell_data in enumerate(row_data):
            worksheet.write(row_num + 1, col_num, cell_data)  # Start writing data from the second row

    workbook.close()
    # Respond with the Excel file
    with open(f"{BASE_DIR+'/media/'+report_meta.report_slug}.xlsx", 'rb') as file:
        response = HttpResponse(file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename={report_meta.report_slug}.xlsx'
    return response
